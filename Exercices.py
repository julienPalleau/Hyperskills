########################################################################################################################
"""
Calculate a remainder
Write a program that calculates the remainder of 10 divided by 3 and prints the result.
"""
# print(10 / 3)
# print(10 % 3)

########################################################################################################################
import csv
import string
from typing import List

import openpyxl

"""
Knowledge verification: Multi-line programs

Write the code that will print all days of the week on separate lines with an empty line between them. Make sure that
each day starts with a capital letter. Start it with Sunday.
"""
# print("Sunday")
# print()
# print("Monday")
# print()
# print("Tuesday")
# print("")
# print("Wednesday")
# print("")
# print("Thursday")
# print("")
# print("Friday")
# print("")
# print("Saturday")

########################################################################################################################
"""
Program with numbers
The first digit of a two-digit number

Write a program that takes a two-digit integer as an input and prints its first digit (i.e., the number of tens).
"""
# number = input()
# print(int(number[0]))

########################################################################################################################
"""
Program with numbers
Calculating an expression

Write a program that reads an integer value n from the standard input and prints the result of the expression:
((n + 1) * n + 2) * n + 3
"""
# n = int(input())
# print(((n + 1) * n + 2) * n + 3)

########################################################################################################################
"""
Program with numbers
Calculating S V P

Ask the user about parameters of a rectangular parallelepiped (3 integers representing the length, width, and height)
and print the sum of edge lengths, its total surface area, and the volume, one answer per line.

Sum of lengths of all edges:
s=4(a+b+c)s = 4(a + b + c)s=4(a+b+c)

Surface area:
S=2(ab+bc+ac)S = 2(ab + bc + ac)S=2(ab+bc+ac)

Volume:
V=abcV = abcV=abc
"""
# a = int(input())
# b = int(input())
# c = int(input())
# s = 4 * (a + b + c)
# S = 2 * (a * b + b * c + a * c)
# V = a * b * c
# print(s)
# print(S)
# print(V)

########################################################################################################################
"""
Program with numbers
Desks

Wow! This problem is kind of tricky. If you're ready to put your thinking cap on, brace yourself and good luck!
Otherwise, you can skip it for now and return any time later.

A school has decided to create three new math groups and equip three classrooms for them with new desks. Your task is to
calculate the minimum number of desks to be purchased. To do so, you'll need the following information:

    The number of students in each of the three groups is known: your program will receive three non-negative integers
    as the input. It is possible that one or more of them will be zero, so you should take it into account.
    Each group will sit in its own classroom. It means that you should calculate the number of desks for each classroom
    separately, and only then add them up!
    At most two students may sit at any desk. You are expected to output the minimum number of desks to buy, so there
    should be as many as possible desks taken by two students rather than one.

Most probably, you'll need operations // and % in your program!
"""
# number_1 = int(input())
# number_2 = int(input())
# number_3 = int(input())
# total = [number_1, number_2, number_3]
# nb_desk = 0
# for number in total:
#     nb_desk += number // 2 + number % 2
# print(nb_desk)

########################################################################################################################
"""
Program with numbers
Savings account

For the given amount of money, calculate the income from this savings account with a 5% interest rate after one year.

Here's the formula for calculating the income: income=amount×interest×years/100

Save the result into the variable income. You DO NOT need to print it.
"""
# amount = 1000
# interest_rate = 5
# years = 1
# # change the next line
# income = 1000 * 5 * 1 / 100
# print(income)

########################################################################################################################
"""
Program with numbers
Difference of times

You will get the values for two moments in time of the same day: when Megan went for a walk, and when it started to
rain. It is known that the first event happened earlier than the second one. Find out how many seconds passed between 
them.

The program gets the input of six integers, each on a separate line. The first three integers represent hours, minutes,
and seconds of the first event, and the rest three integers define similarly the second event. Print the number of
econds between these two moments of time.
"""
# hours_first = int(input())
# minutes_first = int(input())
# seconds_first = int(input())
# hours_second = int(input())
# minutes_second = int(input())
# seconds_second = int(input())
#
# print((hours_second - hours_first) * 3600 + (minutes_second - minutes_first) * 60 + (seconds_second - seconds_first))

########################################################################################################################
"""
Program with numbers
The sum of digits

Given a three-digit integer (i.e., an integer from 100 to 999), find the sum of its digits and print the result.

To get the separate digits of the input integer, make use of % and // (for example, you can get 8 from the number 508
by taking the remainder of the division by 10).
"""
# number = int(input())
# result = []
#
# while number > 0:
#     unit = number % 10
#     number //= 10
#     result.append(unit)
#
# print(sum(result))

########################################################################################################################
"""
Program with numbers
Divide nuts equally between squirrels

N squirrels found K nuts and decided to divide them equally. Determine how many nuts each squirrel will get and print 
the result.
Input data format
There are two positive numbers N and K, each of them is not greater than 10000.
"""
# N = int(input())
# k = int(input())
# print(k // N)

########################################################################################################################
"""
Program with numbers
How many nuts will be left after division

Some squirrels found some nuts and decided to divide them equally. You will get the number of squirrels and the number
of nuts as input on separate lines. Find out how many nuts will be left after each of the squirrels takes an equal
number of nuts. Print the result.
"""
# N = int(input())
# k = int(input())
# print(k % N)

########################################################################################################################
"""
Program with numbers
The last digit of a number
Write a program that reads an integer and outputs its last digit.
"""
# number = list(input())
# print(number[-1])

########################################################################################################################
"""
Program with numbers
Good rest on vacation
Write a program that will help people who are going on vacation. The program should calculate the total required sum
(e.g. $) of money to have a good rest for a given duration.

There are four parameters that have to be considered:

    duration in days
    total food cost per day
    one-way flight cost
    cost of one night in a hotel (the number of nights is equal to the duration of days minus one)

Read integer values of these parameters from the standard input and then print the result.
"""
# duration_days = int(input())
# food_cost_per_day = int(input())
# flight_cost = int(input())  # one way
# hotel_cost = int(input())  # one night
#
# print(f"{duration_days * food_cost_per_day + (duration_days - 1) * hotel_cost + 2 * flight_cost}")
########################################################################################################################
# n = int(input())
# line_list = list(range(1, 2 * n, 2))  # create list of n odd numbers [1, 3, 5, etc.]
# line_max = max(line_list)  # get max width in order to be centered
# for line in line_list:
#     print(('#' * line).center(line_max))

########################################################################################################################
"""
Basic string methods
All in one 
"""
# n = 1
# 'robot'.replace('o', '_', n)  # Replaces n first occurrences of an old substring with a new one.
# "".replace("", 'robot')  # Inserts a substring into the empty line.
# 'robot'.replace('o', "")  # Removes substrings from the line.
# 'robot'.replace("", " ")  # Adds spacing between letters.

########################################################################################################################
"""
Basic string methods
Default
What will be removed using str.lstrip([chars]), str.rstrip([chars]) and str.strip([chars]) methods, if chars is not 
specified?
"""
chars = "test"
print('civic'.strip('vci'))

########################################################################################################################
"""
Basic string methods
Poster artist
Imagine that you design film posters for a living. Write a program that prints each film title in all caps.
"""
# for title in ["the lion king", "MaTRiX"]:
#     print(title.upper())

########################################################################################################################
"""
Basic string methods
Changing the string

Timmy is learning two languages at the same time: English and Python. His teacher told him to work with the following 
string: "London is the capital of Great Britain." Help Timmy! What will be the value of sentence after all the 
operations in the code below?

sentence = "London is the capital of Great Britain."
sentence = sentence.lower()
sentence.upper()
sentence = sentence.replace("a", "x", 2)
sentence.capitalize()
>> london is the cxpitxl of great britain.
"""

########################################################################################################################
"""
Basic string methods
Markdown parsing

Markdown is a special syntax used for text formatting. There are several ways to emphasize words using Markdown:
Typeface 	        Example
Italic text 	    *italics*, _same italics_
Bold text 	        **bold**, __same bold__
Strikethrough text 	~~crossed out~~
Code snippet 	    `code`

Read an input and parse the text so that it doesn't include special symbols mentioned in the table ("*_~`") at its beginning and its end.
Pay attention to the `code` example, ` is not the same as '!

Do not forget to print the result.
"""
# txt = input()
# fixed_txt = txt.strip("*_~`")
# print(fixed_txt)

########################################################################################################################
"""
Basic string methods
Solve a case

Some of the methods we have learned result in similar strings. Match their outputs for the phrase 
"Elementary, my dear Watson".
capitalize()    "Elementary, my dear watson"
title()         "Elementary, My Dear Watson"
swapcase()      "eLEMENTARY, MY DEAR wATSON"
"""

########################################################################################################################
"""
Basic string methods
Preprocessing

Preprocess an input text:

    delete punctuation symbols (commas, periods, exclamation and question marks ,.!?),
    convert all symbols to lowercase.

Then print your text.

Punctuation marks appear not only at the end of the input string. You can use the following way to remove the 
punctuation marks: str.replace("!", ""). Replacing a symbol with an empty string is the same as removing it.
"""
# txt = input()
# print(txt.replace("!", "").replace("?", "").replace(",", "").replace(".", "").lower())

########################################################################################################################
"""
Basic string methods
Latin equivalents
Imagine you have a text-processing program that doesn't work with the ASCII characters. Still, you need to use it and 
don't have time to correct that issue. You've decided to write another program that will replace all the letters with 
ligatures and diacritic marks with their equivalents from the Latin alphabet. The list of replacements is the following:

é -> e
ë -> e
á -> a
å -> aa
œ -> oe
æ -> ae

Complete the code below so that the function normalize (def is a keyword used to declare functions) would take a string 
with diacritics and return the one where all of them have been replaced with the equivalents.
name = input()

def normalize(name):

    # put your code here

    return new_name

print(normalize(name))
"""

# def normalize(name):
#     new_name = ""
#     elemts = {"é": "e", "ë": "e", "á": "a", "å": "aa", "œ": "oe", "æ": "ae"}
#     for n, x in enumerate(name):
#         if x in elemts:
#             new_name += elemts.get(x)
#         else:
#             new_name += x
#
#     return new_name
#
#
# print(normalize("Charlotte Brontë et Lætitia"))


########################################################################################################################
"""
Invoking a function
Young and beautiful

Read the input three times. Each one contains the age of a person – Jack's, Alex's, and Lana's. Find the youngest one
among them and print this person's age.
"""
# jack_age = int(input())
# alex_age = int(input())
# lana_age = int(input())
#
# print(min(jack_age, alex_age, lana_age))

########################################################################################################################
"""
Invoking a function
Hello, world!

The classical introductory exercise, slightly modified. Write a program that takes a string, writes it to the variable
name (this part of code is already written for you), and then prints "Hello, world! Hello, name". Note that there's a
space before the name.
The variable name is already defined.
"""
# name = input()
# print(f"Hello, world! Hello, {name}")

########################################################################################################################
"""
Invoking a function
Longest word

Find the longest word in a pair and print its length.
The variables word1 and word2 are defined for you.
"""
# mot1 = input()
# mot2 = input()
# print(max(len(mot1), len(mot2)))

########################################################################################################################
"""
Create module
Create module

Below there is a custom Python module:

# my_module.py

name = "John"

def printer(x):
    print("Hello,", x)

You have already imported this module into your code:

import my_module

Write a line of the code that will now call the function printer() on the name variable. Mind that both the variable
and the function are defined in my_module.py!
"""
# import my_module
#
# my_module.printer(my_module.name)

########################################################################################################################
"""
Create module
Function main()
Let's say you need a module that:

    stores a string variable n with the value "42".
    defines a main() function that prints the value of n.
    calls the main() function only if it is executed as a script.

Put the lines of this module's code in the correct order and add indents. The first line of the program must be
n = "42". NB: Use the buttons on the left to add indentation (one click equals four spaces).
"""
# n = "42"
#
#
# def main():
#     print("N:", n)
#
#
# if __name__ == "__main__":
#     main()

########################################################################################################################
"""
Errors
Chaos
Look at the code below and eliminate the chaos: the first line should print the resulting number of the calculation and
the second line should print the word mathematics.
print("45/9 + 16*(5 + 8)")
print(mathematics)
"""
# print(45/9 + 16*(5 + 8))
# print("mathematics")

########################################################################################################################
"""
Errors
Locate a problem
The code snippet below needs debugging. Examine it and write down the line number on which the compilation of code will stop.

print( "The quick brown fox" )  # 1
print()                         # 2
print'jumps')                   # 3
pint()                          # 4
print( ' over the lazy dog  )   # 5

answer: 3
"""

########################################################################################################################
"""
Errors
Locate a problem
I am = I'm
I have = I've
I will = I'll
I had / would = I'd
"""
# print("I am = I'm")
# print("I have = I've")
# print("I will = I'll")
# print("I had / would = I'd")

########################################################################################################################
"""
Errors
Locate a problem
Take a look at the following code. Did we declare all variables correctly, according to Python's syntax? Try to fix all
mistakes, so that SyntaxError will not be raised during the initialization of these variables.
hope = "I hope this is actually going to work'
doubt = 'Do I need to double check all lines?'
suggestion = "What if I write it in "double quotes"?
"""
# hope = "I hope this is actually going to work"
# doubt = 'Do I need to double check all lines?'
# suggestion = 'What if I write it in "double quotes"?'

########################################################################################################################
"""
Any and all
Prime numbers

In math, we call a natural number prime if it's greater than 1 and can be divided without any remainder only by 1 and
itself:
2, 3, 5, 7, 11, 13, 17, ...
Create a list of all prime numbers up to 1000 in the code below. Just save this list into a variable prime_numbers,
you don't have to print anything.
Make use of any() or all() to check if a number n leaves a zero remainder when divided by values from 2 to n - 1
(inclusively).
"""
# prime_numbers = [n for n in range(2, 1001) if all(n % x for x in range(2, n))]
# print(prime_numbers)

########################################################################################################################
"""
Any and all
Competition

Today you are taking part in the chess competition. The winner of the competition will get the 'winner' status and the
largest amount of money if they win all the games. Much is at stake!
The results are stored in a list. Fill the blanks in the code below and figure out how much money you won.
You DON'T need to print the answer.
check = ...([True, True, 1, 1, True])

if ...:
    status = 'winner'
else:
    status = 'loser'

if status == ...:
    winning_sum = 100
else:
    winning_sum = 10
"""
# check = all([True, True, 1, 1, True])
#
# if check:
#     status = 'winner'
# else:
#     status = 'loser'
# print(status)
#
# if status == 'winner':
#     winning_sum = 100
# else:
#     winning_sum = 10
# print(winning_sum)

########################################################################################################################
"""
Any and all
Heads or Tails

We have tossed a coin 6 times and saved the results in a list called heads_or_tails. The values are integers: 1 stands
for a head, while 0 denotes a tail.
Add some code to find out whether the list has any heads. Do not print the variable check, just store the result in it.
# Fingers crossed
check =  # are there any heads in the list heads_or_tails?
"""
# heads_or_tails = [0, 0, 0 ,0, 0, 0, 0, 0]
# check = any(heads_or_tails)
# print(check)

########################################################################################################################
"""
Any and all
Lottery

Imagine that you have bought a bunch of lottery tickets and wrote down their numbers into the list. Now, it's time to
figure out whether you have a winning ticket.
You know that all winning tickets are no less than 44. Fill the empty fields in the code (these ones ...) to check if 
you have at least one winning ticket.
You DON'T need to print the answer.
# As luck would have it
tickets = [11, 22, 33, 44, 55]
winning_tickets = [i >= ... for i in ...]
tickets_bool = ...(winning_tickets)
"""
# # As luck would have it
# tickets = [11, 22, 33, 44, 55]
# winning_tickets = [i >= 44 for i in tickets]
# tickets_bool = any(winning_tickets)
# print(tickets_bool)
# print(any("the object"))

########################################################################################################################
"""
Exception handling
Implementing logic

Advanced input() handling is used to read input directly into several variables, e. g.,
name, surname = input().split(), but a user still can enter more or fewer words.
In this task, you have to make sure that the user entered the necessary amount of words and...
    If there are more or fewer words in the input, print an error: "You need to enter exactly 2 words. Try again!"
    If everything's good, greet the user personally.
"""
# try:
#     name, surname = input().split()
# except Exception:
#     print("You need to enter exactly 2 words. Try again!")
# else:
#     print(f"Welcome to our party, {name} {surname}")

########################################################################################################################
"""
Exception handling
Zero

You know how to catch the built-in exceptions. Right now, try to read two numbers (a, b) and find the result of their
division. Your main task it to catch the ZeroDivisionError. If there's an error, print the following message:
The Error!. Otherwise, print the result of the division.
a = int(input())
b = int(input())
"""
# a = int(input())
# b = int(input())
#
# try:
#     print(a / b)
# except ZeroDivisionError:
#     print("The Error!")

########################################################################################################################
"""
Exception handling
Chatty bot

The original exercise consist in re-ordering the code below:
finally:
print("Hope to see you soon!")
print("What a beautiful name you have!")
except NameError:
try:
print("Hello, stranger!")
print("Hello,, name")

"""
# try:
#     print("Hello,, name")
# except NameError:
#     print("Hello, stranger!")
# else:
#     print("What a beautiful name you have!")
# finally:
#     print("Hope to see you soon!")

########################################################################################################################
"""
Exception handling
Modeling situation

Your program will have access to the exception_test() function, which can throw exceptions.
You need to write code that runs this function, then catches ArithmeticError, AssertionError, ZeroDivisionError
exceptions and prints the name of the caught exception. Mind the hierarchy of exceptions.
An example solution that you should send for review:
try:
    exception_test()
except Exception:
    print("Exception")
except BaseException:
    print("BaseException")
"""

# def exception_test():
#     return 3 / 0
#
#
# try:
#     exception_test()
# except ZeroDivisionError:
#     print("ZeroDivisionError")
# except ArithmeticError:
#     print("ArithmeticError")
# except AssertionError:
#     print("AssertionError")

########################################################################################################################

"""
Split and join
Find positions

Write a program that reads a sequence of numbers from the first line and the number x from the second line. Then it 
should output all positions of x in the numerical sequence.
The position count starts from 0. In case x is not in the sequence, print the line "not found" (quotes omitted,
lowercased).
Positions should be displayed in one line, in ascending order of the value.
"""

# input_string = input()
# my_list = input_string.split()
# str_to_int = list(map(int, my_list))
# number = int(input())
# result = []
# test = False
#
# for ind, num in enumerate(str_to_int):
#     if num == number:
#         result.append(ind)
#         test = True
# res = []
# if test:
#     [res.append(ind) for ind, num in enumerate(str_to_int) if num == number]
#     [print(line, end=" ") for line in res]
# else:
#     print("not found")

########################################################################################################################
"""
Split and join 
Triangle 

Draw a triangle of a given height, as in the example.
 Sample Input 1:
3

Sample Output 1:
  #  
 ### 
#####
"""
# # Solution 1:
# number = int(input())
# element = "#"
# for i in range(number+2):
#     if i % 2 == 1:
#         print((element * i).center(number+2))

# Solution 2:
# height = int(input())
# width = height + (height - 1)
# block = "#"
#
# for _ in range(height):
#     print(block.center(width))
#     block += "##"

########################################################################################################################
"""
Split and join
What day is It?

Read a date from the input given in one of the following formats: YYYY-MM-DD or YYY-MM-DD. Print the year, month and 
day on separate lines.
 Sample Input 1:

2020-04-30

Sample Output 1:

2020
04
30
"""
# date = input()
# print("\n".join(date.split("-")))


########################################################################################################################
"""
- Split and join -
String tricks

Examine the code and fix the mistakes so that the join() method works.
random_numbers = [1, 22, 333, 4444, 55555]
print("\n".join(random_numbers))
"""
# random_numbers = [1, 22, 333, 4444, 55555]
# print("\n".join([str(i) for i in random_numbers]))

########################################################################################################################
"""
- Split and join -
Prepositional genitive

Advanced input() handling is used to read input directly into several variables, for example:
x, y = input().split()
"""
# x, y = input().split()
# print(f"{x} of {y}")

########################################################################################################################
"""
- Split and join -
Straight A's
Write a program that calculates the proportion of students who received grade A.
A five-point rating system with grades A, B, C, D, F is used.
Input format:
A string with students' marks separated by space. At least one mark will be present.
Output format:
A fractional number with exactly two decimal places.
To format the decimal places use the round(number, places) function.
"""
# notes = input()
# notes = notes.split()
# print(round(notes.count("A") / len(notes), 2))

########################################################################################################################
"""
- Split and join -
Flip
Let's try to reverse a numeric sequence. Read it from the input and print its numbers in reverse order separated by
spaces.
"""
# # solution 1
# sequence = input().split()
# [print(num, end=" ") for num in sequence[::-1]]
# print()
#
# # solution 2
# sequence = input().split()
# print(" ".join(sequence[::-1]))

########################################################################################################################
"""
- Split and join -
Find words
Find all words that end in "s" and print them together separated by an underscore.
After such words there will be no punctuation marks, you do not need to worry about that.
"""
# # Solution 1
# sentence = input().split()
# result = ""
# for word in sentence:
#     if word[-1] == "s":
#         result += word + "_"
#
# print(result[:len(result) - 1:])
#
# # Solution 2
# sentence = input().split()
# new = []
# for word in sentence:
#     if word.endswith('s'):
#         new.append(word)
# print("_".join(new))

########################################################################################################################
"""
- Split and join -
mixedCase

Convert a text into lowerCamelCase, or mixedCase.
The input format:
A lowercased word or several words separated by spaces.
The output format:
Print this text stylized as mixedCase, that is, the first word should be in lowercase and all other words – capitalized.
"""
# # Solution 1
# text = input().split()
# result = text[0]
# i = 0
# for word in text:
#     if i != 0:
#         result += word.capitalize()
#     i += 1
# print(result)
#
# Solution 2
# word = input().title().split()
# word[0] = word[0].lower()
# print("".join(word))

########################################################################################################################
"""
- Split and join -
Spellchecker 

Write a spellchecker that tells you which words in the sentence are spelled incorrectly. Use the dictionary in the code 
below.
The input format:
A sentence. All words are in the lowercase.
The output format:
All incorrectly spelled words in the order of their appearance in the sentence. If all words are spelled correctly, 
print OK.
dictionary = ['all', 'an', 'and', 'as', 'closely', 'correct', 'equivocal',
              'examine', 'indication', 'is', 'means', 'minutely', 'or', 'scrutinize',
              'sign', 'the', 'to', 'uncertain']
"""
# dictionary = ['all', 'an', 'and', 'as', 'closely', 'correct', 'equivocal',
#               'examine', 'indication', 'is', 'means', 'minutely', 'or', 'scrutinize',
#               'sign', 'the', 'to', 'uncertain']
#
# test = True
# sentence = input().split()
# for word in sentence:
#     if word in dictionary:
#         pass
#     else:
#         print(word)
#         test = False
#
# if test:
#     print("OK")

########################################################################################################################
"""
- Split and join -
Fix the mistakes

The code below is supposed to find all website addresses (https://, http://, www.) in the input text. However, it is 
not finished. Complete the code and print all the addresses in the order in which they appear in the text, each on a 
new line. Mind that str.lower() can help you to convert all characters of the string to the lower case.
"""
# # Solution 1
# import re
# text = input()
# words = text.split()
# for word in words:
#     r1 = re.search(r"([w+W+]{3}.*)", word)
#     if r1:
#         print(r1.group(0))

# # Solution 2
# text = input()
# words = text.split()
# for word in words:
#     if word.lower().startswith("www.") or word.lower().startswith("https://") or word.lower().startswith("http://"):
#         print(word)

# # Solution 3
# text = input().split()
# addresses = ("https://", "http://", "www.")
#
# for word in text:
#     if word.lower().startswith(addresses):
#         print(word)

########################################################################################################################
"""
- Split and join -
CapWords

In Python, the names of classes (you'll learn about them in detail later) follow the CapWords convention. Let's convert 
the input phrase accordingly by capitalizing all words and spelling them without underscores in-between.
The input format:
A word or phrase, with words separated by underscores, like function and variable names in Python.
You might want to change the case of letters since they are not necessarily lowercased.
The output format:
The name written in the CapWords fashion.
"""
# sentence = input()
# result = sentence.split("_")
# res = ""
# for word in result:
#     res += word.capitalize()
# print(res)

########################################################################################################################
"""
The Shutil module
Which error? 
Suppose you are trying to move a directory to another folder. However, you are getting an error because the directory 
you are moving is in the target folder. What error will be raised?
Select one option from the list
FileNotFoundError
FileExistsError
KeyError
>>> shutil.Error
"""
# # Python program to explain shutil.copytree() method
#
# # importing os module
# import os
#
# # importing shutil module
# import shutil
#
# # path
# path = 'C:/Users/MOTTIER LUCIE/Documents/GitHub/Hyperskills'
#
# # List files and directories
# # in 'C:/Users / Rajnish / Desktop / GeeksforGeeks'
# print("Before copying file:")
# print(os.listdir(path))
#
# # Source path
# src = 'C:/Users/MOTTIER LUCIE/Documents/GitHub/Hyperskills/source'
# print("Within source path:")
# print(os.listdir(src))
# #
# # Destination path
# dest = 'C:/Users/MOTTIER LUCIE/Documents/GitHub/Hyperskills/destination'
# print("Within destination path:")
# print(os.listdir(dest))
# #
# # # Copy the content of
# # # source to destination
# destination = shutil.copytree(src, dest, dirs_exist_ok=True)
# print("After copying file:")
# print(os.listdir(dest))
# #
# # # Print path of newly
# # # created file
# # print("Destination path:", destination)

########################################################################################################################
"""
The Shutil module
Matching functions 

shutil.copy()           Copies the content of the file to the destination file or directory
shutil.rmtree()         Delete a directory
shutil.move()           Moves a file or directory to the destination
shutil.make_archive()   Compress a file in the given format
shutil.which()          Finds the path to the executable file
"""

########################################################################################################################
"""
The Shutil module
How to keep the metadata?

Sam wants to copy a file on his computer. He wants to keep the metadata. Which function would you advise to use?
shutil.copy()
x shutil.copy2()
shutil.copyfile()
shutil.which()

"""

########################################################################################################################
"""
The Shutil module
Find an executable file 

Recently, you have decided to learn C++ and downloaded it to your computer. You want to make sure the file has been 
downloaded successfully. In the text field, write the function that will find the location of C++. You do not need to 
print the result.
"""
# print(shutil.which("python"))

########################################################################################################################
"""
The Shutil module
Be in order

To organize the folders, the football folder needs to be copied into the sports folder. We keep the path of the 
football directory in the path_football variable; the sports directory in the path_sports variable. Which function 
can help us with this operation? In the text field, state the required function.
"""
# shutil.copytree(path_football, path_sports)

########################################################################################################################
"""
The Shutil module
File compression 

You need to compress the letters file. Its location is stored in the path variable. How do we perform the compression 
operation? Use comp_letters for the new name of the file and any standard type of archive. Don't save it into a 
variable, just type the command to create such an archive.
"""
# import shutil
#
# directory = "/home/user/Desktop/Algorithms"
# shutil.make_archive("comp_letters", "bztar", path)

########################################################################################################################
import glob

"""
Glob module
Help a friend

Your friend is a language teacher that has the my_dir directory with all kinds of files. Help him find all txt files 
which have a1, a2 or b1 in their names. Use the glob module for that and take a look at what your friend has written:
glob.glob('my_dir\\*[a-c]1*.txt')
There's a mistake in this search. Which files does it return?
Select one or more options from the list
x   my_dir\\excercises b1.txt
    my_dir\\some b2 texts.txt
x   my_dir\\new c1 materials.txt
    my_dir\\a2 texts.txt
x   my_dir\\a1_texts.txt
"""

########################################################################################################################
"""
https://docs.python.org/3/howto/regex.html
Glob module
How to open all.txt files?

Imagine you have a folder with files in it. Some of them have the txt extension, and others have the HTML extension. 
Write a program that allows you to find the ones which have the txt-extension. Sort them in alphabetical order and 
print the final list.

You can use the sorted(your_list_name) function for sorting. Consider that you are already inside the directory.
"""
# import os
#
# print(os.getcwd())
# print(glob.glob('*'))
# print(glob.glob('[__ a-z]*'))

########################################################################################################################
"""
Glob module
Wildcard meanings

Let's refresh our memory about what wildcards mean. Match them to their meanings.
*           Matches 0 or more characters.
?           Matches only 1 character.
[ - ]       Specifies a range of characters.
[ ]         Matches 1 character from the given sequence.
[! ]        Matches any character not in the provided sequence.
"""

########################################################################################################################
"""
Glob module
Simple pathnames
Let's match pathnames to the files they will match in the glob.glob() method.

glob.glob('my_dir\\*.txt')      All .txt files.   
glob.glob('my_dir\\?.*')        Filenames that consist of one character.
glob.glob('my_dir\\?.txt')      .txt files with one character names.
glob.glob('my_dir\\?')          Subdirectories with one character names.
"""
# import os
#
# print(os.getcwd())
# print(glob.glob('?.*'))

########################################################################################################################
"""
Glob module
2014 — 2018 reports
Let's say you have some txt files with reports in my_dir. Their names are report_2012.txt, report_2013.txt and so on, 
all the way up to 2021. You need to find the reports from 2014 to 2018. Which of the following lines of code could 
you use to do that? 

    glob.glob('my_dir\\report_20[12-19].txt')
    glob.glob('my_dir\\report_201?.txt')
    glob.glob('my_dir\\report_*.txt')
x   glob.glob('my_dir\\report_201[4-8].txt')
"""

########################################################################################################################
"""
Glob module
iglob
What does glob.iglob() return? 
x   A generator.
    A dictionary.
    A string.
    A list.
"""

########################################################################################################################
"""
Glob module
Recursive flag 
What does the following method return?
glob.glob('my_dir\\**', recursive=True)

    A list of all files inside subdirectories.
    A list of subdirectories.
x   A list of all files, subdirectories, and files inside the subdirectories.
    A list of all files and subdirectories.
"""

########################################################################################################################
"""
Glob module
Advanced pathnames
Match search patterns to what you will find.

glob.glob('my_dir\\b*[!bf].*')  Filenames starting with b but not ending with b or f.
glob.glob('my_dir\\[bf]*.*')    Filenames starting with either b or f.
glob.glob('my_dir\\[b-f]*.*')   Filenames starting with b, c, d, e or f.
glob.glob('my_dir\\*[!b]')      Subdirectories with names which do not end in b.
"""

########################################################################################################################
"""
Glob module
A generator

Let's say you want to create a generator yielding the names of all csv files in my_dir and print them. Put the lines of 
code in the correct order. Don't forget about the indentation.
    import glob
    generator = glob.iglob("my_dir\\*.csv")
    for item in generator:
    print(item)
"""

########################################################################################################################
"""
Glob module
A simple asterisk 

X   A list of all files and subdirectories in my_dir.
    a list of all .txt files in my_dir.
    A list of all files in my_dir.
    A list of all subdirectories in my_dir.
    It raises an error.
"""

########################################################################################################################
"""
Openpyxl
Theory
"""
# # First steps
# from openpyxl import Workbook
# book = Workbook()
# new_sheet = book.active
# print(type(new_sheet))
#
# new_sheet['A1'] = 'cat'
# new_sheet['A2'] = 'dog'
# new_sheet['A3'] = 'tiger'
# new_sheet['A4'] = 'hamster'
# new_sheet.cell(row=5, column=1).value = 'sparow'
#
# book.save("New_file.xlsx")
#
# # Opening a file to read/write it
# workbook = openpyxl.load_workbook('New_file.xlsx')
#
# # Reading Files
# all_sheets_names = workbook.sheetnames
# sheet = workbook[all_sheets_names[0]]
# print(sheet['A1'].value)
# print(sheet['A2'].value)
# print(sheet['A3'].value)
# print(sheet['A4'].value)
# print(sheet['A5'].value)
#
# # Writting to a file
# workbook.create_sheet(title='for_testing')
# write_sheet = workbook['for_testing']
#
# write_sheet['A1'] = 'Tom'
# write_sheet['B1'] = 'Black'
# write_sheet['C1'] = 'The Hague'
# write_sheet['D1'] = 'We will discuss later'
# write_sheet['E1'] = 5
#
# # Reading Files
# sheet = workbook['for_testing']
# print(sheet['A1'].value)
# print(sheet['B1'].value)
# print(sheet['C1'].value)
# print(sheet['D1'].value)
# print(sheet['E1'].value)
#
# # Filtering entries
# print()
# workbook = openpyxl.load_workbook('sample-xls-file-for-testing.xlsx')
#
# sheet = workbook['Sheet1']
#
# rows = sheet.max_row
# cols = sheet.max_column
# print(f"Number max rows: {rows}, columns: {cols}")
#
# # To print all entries
# print()
# # for row in range(1, rows + 1):
# #     string = ''
# #     for column in range(1, cols +1):
# #         value = sheet.cell(row=row, column=column).value
# #         string = string + str(value) + ', '
# #     string = string[:-2]  # we remove the last appended comma and space
# #     print(string)
#
# # Let's take a step further and print only those entries that contain Germany:
# for row in range(1, rows + 1):
#     string = ''
#     for column in range(1, cols + 1):
#         value = sheet.cell(row=row, column=column).value
#         string = string + str(value) + ', '
#     string = string[:-2]
#     check = sheet.cell(row=row, column=2)
#     check = str(check.value)
#     if check == 'Germany':
#         print(string)

########################################################################################################################
"""
Openpyxl
Creating a file

You need to create a new Excel file and write down the name of your new classmate. You have the code but, 
unfortunately, the lines are scrambled! Place them in the right order. 
from openpyxl import Workbook
new_book = Workbook()
new_sheet = new_book.active
new_sheet['A1'] = 'Giuliana'
new_book.save("The_name_of_my_friend.xlsx")
"""

########################################################################################################################
"""
Openpyxl
Opening a file 

    workbook.loading_workbook()
    workbook.upload_workbook()
    workbook.download_workbook()
X   workbook.load_workbook()
"""

########################################################################################################################
"""
Openpyxl
How to process?

If our Excel file contains thousands of data entries to be processed, we can use if-statements and 
____________ together.
X   for-loops
    nothing more
    functions
    other libraries
"""

########################################################################################################################
"""
Openpyxl
Maximum rows

You are the owner of a big company. There is a very big file with tons of information about your customers on your 
computer. You need to find out the maximum number of rows in an existing file. What attributes can be used for that?

sheet.max_row
"""

########################################################################################################################
"""
Openpyxl
Line description

Match each line with what it does.
from openpyxl import Workbook   We import the Workbook class.
book = Workbook()               We save a new Workbook in a variable.
new_sheet = book.active         We create a working environment.
"""

########################################################################################################################
"""
Openpyxl
Variable type

You created a new sheet. This sheet was stored in a variable. What is the type of this variable?
X   <class 'openpyxl.worksheet.worksheet.Worksheet'>
    <class 'openpyxl.workbook.workbook.Workbook'>
    <class 'openpyxl.workbook.Workbook'>
    <class 'openpyxl.worksheet.Worksheet'>
"""

########################################################################################################################
"""
Openpyxl
Writing to a cell

Imagine that you have created a program that writes down city names to an Excel file. One of the code lines reads the 
following: sheet_one.cell(row=3, column=1).value = 'La Habana'. Use the shorter notation and rewrite the line to save 
the city in the same cell.

sheet_one['A3'] = 'La Habana'
"""

########################################################################################################################
"""
Openpyxl
Other libraries

In this topic, we asked you to have a look at other libraries that can be alternatives for openpyxl on the special page.
Please, choose all the right alternatives below.
X   xlsxwriter
X   xlrd
    xltw
    pylightlx
X   pyxlsb
"""

########################################################################################################################
"""
Working with PDF in Python
https://pythonguides.com/create-and-modify-pdf-file-in-python/

Extract the metadata

We can extract the metadata from PDF files using the pdf.getDocumentInfo() method. Suppose we store the metadata in the 
information variable. Let's explore the Author, Producer, and Title from dummy.pdf.

# get the author, producer and title from the 'information' dict
print(information.author)
print(information.producer)
print(information.title)
"""

########################################################################################################################
"""
Working with PDF in Python
Add pages to PDF

Suppose you want to create a new pdf file and add a lot of cells. Let's organize the following lines!
Remember that first of all, you need a page. After that, you need to set a font before writing in the document.
"""
# from fpdf import FPDF
#
# pdf = FPDF(orientation="P", unit="mm", format="Letter")
# pdf.add_page()
# pdf.set_font(family="times", style="I", size=10)
#
# for i in range(5):
#     pdf.cell(w=0, h=20, txt=f"This is the line {i}", border=0, ln=1, align="L")
#
# pdf.output("file.pdf")

########################################################################################################################
"""
Working with PDF in Python
Create a small PDF file

Suppose you want to create a new PDF document in the landscape mode, with mm units and the A4 format. Use the FPDF 
class to write an object called dummy_pdf with the mentioned attributes. Please, follow the described order and use 
double quotation marks.
"""
# from fpdf import FPDF, XPos, YPos
#
# dummy_pdf = FPDF(orientation="L", unit="mm", format="A4")
# dummy_pdf.add_page()
# dummy_pdf.set_font(family='times', style='I', size=14)
# dummy_pdf.cell(w=1, txt="hello world!")
# dummy_pdf.output("dummy.pdf")

########################################################################################################################
"""
Working with PDF in Python
Rotate a page

We can rotate pages in a PDF using the library PyPDF4. Select all valid methods to rotate pages:
X   rotateClockwise(degrees)
    rotateLeftClockwise(degrees)
    rotate_pages_left(degrees)
X   rotateCounterClockwise(degrees)
    rotate_pages_right(degrees)
"""

########################################################################################################################
"""
Working with PDF in Python
Create a PDF file

Organize the following lines to create a sample PDF document. Remember that before printing text or a cell, it is 
mandatory to select a font, otherwise, the document would be invalid. 

from fpdf import FPDF
pdf = FPDF()
pdf.add_page()
pdf.set_font('times', 'B', 16)
pdf.cell(40, 10, 'This is my first document')
pdf.output('dummy.pdf')
"""

########################################################################################################################
"""
Working with PDF in Python
PDF functions

Match the code lines with their function:
Read a PDF file                         PdfFileReader()
Get the number of pages                 getNumPages()
Write a PDF file                        PdfFileWriter()
Create a new PDF file                   FPDF("P", "cm", "Letter")
Extract the metadata from a PDF file    getDocumentInfo()
Add a new page                          add_page()
"""

########################################################################################################################
"""
Working with PDF in Python
Extract the metadata

What is the function of the PdfFileReader class that allows you to extract the metadata from a PDF file?
    getMetaData()
    extract_metadata()
X   getDocumentInfo()
    get_document_info()

"""

########################################################################################################################
"""
Working with PDF in Python
Add the markdown

Fpdf2 supports basic Markdown-like styling: **bold**, __italics__, --underlined--. We only need add an optional 
markdown=True parameter to the cell method.
Let's organize the following lines to add some markdown to a sample cell.

from fpdf import FPDF
pdf = fpdf.FPDF()
pdf.add_page()
pdf.set_font("Times", size=60)
pdf.cell(txt="**Hello** __word__ --python--", markdown=True)
pdf.output("sample.pdf")
"""

########################################################################################################################
"""
Working with PDF in Python
Add HTML to a PDF file 

The fpdf2 library supports basic rendering from HTML. Explore its documentation and select the method used to compose 
basic HTML on a page.

    set_format()
X   write_html()
    writeHtml()
    set_html()
"""

########################################################################################################################
"""
Files in Python
Which one?

What mode should be used for updating a file (reading and writing)? It should be one symbol, not a combination.
>> +
"""

########################################################################################################################
"""
Files in Python
Missing

Below you can see a code snippet for working with a file. However, a crucial line of code is missing. Add the missing 
line to the code. 
file = open('test_file.txt', 'w')
file.write('This line will be in the file!')
"""
# file = open('test_file.txt', 'w')
# file.write('This line will be in the file!')
# file.close()

########################################################################################################################
"""
Files in Python
The parameter

What parameter regulates how we want to open our file and what for?
>> mode
"""

########################################################################################################################
"""
Files in Python
Mojibake

Mary had a file some_text.txt that she wanted to open using this code:
file = open('some_text.txt', 'r', encoding='cp1251')
When she tried to print the content of the file, she got this:
С†РёР°РЅРјР°РґР¶РµРЅС‚Р°Р¶РµР»С‚С‹Р№РєР»СЋС‡РµРІРѕР№ С†РІРµС‚
Which parameter of the open() function did she specify incorrectly? Below, write the parameter, for example, name or 
mode.
>>> encoding
"""

########################################################################################################################
"""
Files in Python
Mode

Which mode of the open() function allows you to open the file for writing in binary without overwriting the existing 
contents of the file?
    a+
X   ab
    at
    rb
    w+
    wb
"""

########################################################################################################################
"""
Files in Python
Error

Alice has a file called "books.txt" with all the books that she needs to read for her class. Her professor added 
another book to the list, so Alice wants to add a new line to the file. Below you can see her code.
However, Alice made a mistake and got a result which was absolutely not what she had expected. Can you fix her mistake?
Don't worry about closing the file here, it will be done under the hood.

# find a mistake below
file = open('books.txt', 'w', encoding='utf-8')

# closing and other operations are done below
"""
# # find a mistake below
# file = open('books.txt', 'a', encoding='utf-8')
#
# # closing and other operations are done below

########################################################################################################################
"""
Files in Python
Correct files

Which files below have been opened correctly?
X   file1 = open('file1.txt')
    file2 = open('file2.txt', 'rb', encoding='utf-8')
    file3 = open('file3.txt', 'abt')
X    file4 = open('file4.txt', 'w+', encoding='utf-16')
    file5 = open()
"""

########################################################################################################################
"""
Files in Python
Open the file

To complete this task, follow the instructions.

    Create a new file "stars.txt".
    Open it for writing as a text.
    Use the UTF-16 encoding.
    Name the variable test_file.
    Don't forget to close the file!
"""
# test_file = open('stars.txt', 'wt', encoding='utf-16')
# test_file.close()

########################################################################################################################
"""
Files in Python
Inappropriate combination

Some modes cannot be combined with each other: thus, only one of the options 'w', 'r', and 'a' must be specified, we 
can't open a file with 'ar' mode. Similarly, we must choose either 'b' or 't', the file can't be opened both in the 
text and binary modes.

What modes can't be combined?
X   a and r
    r and t
    w and b
X   b and t
"""

########################################################################################################################
"""
Files in Python
Which mode?

Suppose, you want to open some file for both reading and writing. You don't want to lose the contents of your file and 
you also don't want to get an error if by some chance the file doesn't exist. Which mode should you choose?

>> a+
"""

########################################################################################################################
"""
Reading files
Summer

Please note: for now, when you go to the tab IDE, you can solve this problem only via Pycharm. We are sorry for the 
inconvenience, we will fix this as soon as we can. If you prefer to solve the problem via a different IDE, you should 
open it manually and then copy the answer to the website.

To solve this problem, you will need a file – download it below. Your task is to count how many times the word summer 
appears in the file. It contains only one word per line, all of them are lowercased. Mind that if summer is a part of 
some other word (e.g. summertime) you should NOT count it. We only need summers themselves! Also, you don't need to 
enter any code here, just write and run it on your machine, and then enter the result you get.
"""
# file = open('hyperskill-dataset-71758223.txt', 'r')
# count_summer = 0
# for line in file:
#     print(line)
#     if line == 'summer\n':
#         count_summer += 1
# file.close()
# print(f"nombre de line summer: {count_summer}")

########################################################################################################################
"""
Reading files
Test file

Suppose we have a file test.txt that looks like this:
This
is
a
test
file

Regard \n as the newline escape sequence.

We want to read 10 bytes from this file:

file = open("test.txt", 'r')
print(file.read(10))
file.close()

What will be the output of this code?

>>>
This
is
a
"""

########################################################################################################################
"""
Reading files
Methods for reading

read()      Reads the file as a whole.
readline()  Reads the file one line at a time.
readlines() Reads the file as a list of lines.
"""

########################################################################################################################
"""
Reading files
Newline escape sequences

We know that lines in files are separated by the special newline escape sequences. Find all such sequences below.
    \b
X   \n
    \t\b
    \t
X   \r\n
X   \r
"""

########################################################################################################################
"""
Reading files
Acronym

You are given the file test.txt. Read it and print the first letter from each line.
Don't forget to close the file!
"""
# file = open('hyperskill-dataset-71758223.txt', 'r')
#
# for line in file:
#     print(line[0])
# file.close()

########################################################################################################################
"""
Reading files
First

There is a file called test_file.txt. Read this file and print the first line. Specify encoding='utf-16'.
Don't forget to close the file!
"""
# file = open('test_file.txt', 'r', encoding='utf-16')
# print(file.readline())
# file.close()

########################################################################################################################
"""
Reading files
War and Peace

Suppose you have a file that contains the entire text of "War and Peace" by Leo Tolstoy. In this file, each sentence of 
the novel is on a new line.

You want to write a program that preprocesses the text (e.g. splits sentences into words, and deletes punctuation 
marks). What is the optimal way to read this file?

X   for loop
    readlines()
    read()
    readline()
"""

########################################################################################################################
"""
Reading files
Line number
"""

# file = open('hyperskill-dataset-71758223.txt', 'r')
# Solution 1
# print(file.read().count('\n'))

# Solution2
# print(len(file.readlines()))

# ATTENTION la solution 1 et 2 ne fonctionnent pas si elles sont executes toutes les 2. Il faut enlever le commentaire
# que sur une ligne à la fois. C'est à cause du curseur qui ne revient pas a sa position initiale

# Solution 3
# counter = 0
# for _ in file:
#     counter += 1
# print(counter)
# file.close()

########################################################################################################################
"""
Reading files
Sum

You have the file sums.txt. It has several lines, each of them containing two positive numbers separated by a 
whitespace.
For example:
9 61
15 47
2 1
Your program should read this file and print the sum of numbers on each line. So, if the file has n lines, you should 
print n sums, each on a separate line.
Don't forget to close the file!
"""
# file = open('sums.txt', 'r')
# for line in file.readlines():
#     a, b = line.split()
#     print(int(a) + int(b))
# file.close()

########################################################################################################################
"""
Reading files
Seasons

There's a file seasons.txt that looks like this:
Spring
Summer
Autumn
Winter
We implemented the following code:

seasons_file = open('seasons.txt', 'r', encoding='utf-8')
seasons = seasons_file.readlines()
favorite_season = seasons[2]
seasons_file.close()

What will be the value of the variable favorite_season? Write it below without quotation marks. Remember that the 
default newline character is '\n'.
answer: Autumn\n
"""

########################################################################################################################
"""
Writing files
What's the difference?

What is the difference between write()and writelines()?
"""

########################################################################################################################
"""
Writing files
Solar system

Create a file planets.txt and write the names of the Solar system planets there, each on a new line. In total, the file 
should contain 8 lines with the following planets: Mercury, Venus, Earth, Mars, Jupiter, Saturn, Uranus, and Neptune.
When opening the file, specify encoding='utf-8'.
"""
# # Solution 1
# file = open('planets.txt', 'w', encoding='utf-8')
# for name in ("Mercury", "Venus", "Earth", "Mars", "Jupiter", "Saturn", "Uranus", "Neptune"):
#     file.write(f'{name}\n')
# file.close()
#
# # Solution 2
# planets = ['Mercury\n', 'Venus\n', 'Earth\n', 'Mars\n', 'Jupiter\n', 'Saturn\n', 'Uranus\n', 'Neptune']
# file = open('planets2.txt ', 'w', encoding='utf-8')
# file.writelines(planets)
# file.close()

########################################################################################################################
"""
Writing files
1 to 10

We have a file numbers.txt that contains a single line with the number 1. Suppose, we run the following code:
file = open('numbers.txt', 'a')

for i in range(2, 11):
    file.write(str(i) + ' ')

file.close()

How many lines will numbers.txt contain after this?
"""
# file = open('numbers.txt', 'a')
#
# for i in range(2, 11):
#     file.write(str(i) + ' ')
#
# file.close()

########################################################################################################################
"""
Writing files
Sort the code

In this task, you have to sort the lines of the code given below. Mind that first we need to create a list of animals.
animals = ['cat\n', 'dog\n', 'hamster\n']
animal_file = open('animals.txt', 'w', encoding='utf-8')
animal_file.writelines(animals)
animal_file.close()
"""

########################################################################################################################
"""
Writing files
Songs

Suppose we have a file songs.txt that contains a single line: Space Oddity. We have implemented the following code:

f = open('songs.txt', 'w', encoding='utf-8')
f.write('Life on Mars')
f.close()

What will be the contents of songs.txt after this?
    Space Oddity Life on Mars
    Space Oddity
    Life on Mars
    Space Oddity
X   Life on Mars
"""

########################################################################################################################
"""
Writing files
Input

Read a line from input and write it to the file input.txt.
"""
# text = input()
# f = open('input.txt', 'w')
# f.write(text)
# f.close()

########################################################################################################################
"""
Writing files
Travel

There's a file countries.txt that contains a list of the countries Kate has visited. For example:

France
China
Brazil

During her last trip, she visited Turkey for the first time. Add Turkey to the countries.txt file.

Kate will not give up traveling, so it would be reasonable to add a newline character at the end.

Don't forget to close the file!
"""
# # animals = ['cat\n', 'dog\n', 'hamster\n']
# # animal_file = open('animals.txt', 'w', encoding='utf-8')
# # animal_file.writelines(animals)
# # animal_file.close()
#
# # # Solution 1
# last_country = ['Turkey\n']
# file = open('countries.txt', 'a', encoding='utf-8')
# file.writelines(last_country)
# file.close()
#
# # Solution 2
# last_country = ['Turkey\n']
# file = open('countries2.txt', 'a', encoding='utf-8')
# file.write(last_country[0])
# file.close()

########################################################################################################################
"""
Writing files
CMYK

You have a list of strings color_list and you have implemented the following code to write it to the file CMYK.txt:

color_list = ['cyan', 'magenta', 'yellow', 'key color']

cmyk_file = open('CMYK.txt', 'w', encoding='utf-8')
cmyk_file.writelines(color_list)
cmyk_file.close()

What will be the contents of the file CMYK.txt?
"""
# color_list = ['cyan', 'magenta', 'yellow', 'key color']
#
# cmyk_file = open('CMYK.txt', 'w', encoding='utf-8')
# cmyk_file.writelines(color_list)
# cmyk_file.close()

########################################################################################################################
"""
Writing files
Animals

The file animals.txt has a list of animals, each written on a new line. For example:

rabbit
cat
turtle

Create a new file, animals_new.txt, where those animals are written on a single line and separated by a whitespace.

Don't forget to close all files!
"""
# # read animals.txt
# # and write animals_new.txt
# with open('animals.txt') as file_1:
#     animals = file_1.read().replace('\n', ' ')
#     with open('animals_new.txt', 'w') as file_2:
#         file_2.write(animals)

########################################################################################################################
"""
Context manager
Error 

The code below is supposed to write years from 2010 to 2020 to the file years.txt. However, something's wrong with the code.
Find the mistake and fix it.
with open('years.txt', 'w', encoding='utf-8') as f:
    for i in range(2010, 2020):
        f.write(str(i) + " ")

f.write('2020')

Solution:
with open('years.txt', 'w', encoding='utf-8') as f:
    for i in range(2010, 2020):
        f.write(str(i) + " ")

    f.write('2020')
"""

########################################################################################################################
"""
Context manager
Rewrite

Rewrite the following code using the with keyword:
f = open('test.txt', 'w')
f.write('Tada!')
f.close()

f.close()

Please open a context manager with the name f.
"""
# with open('test.txt', 'w') as f:
#     f.write('Tada!')

########################################################################################################################
"""
Context manager
The purpose

What is the main purpose of context managers?
    Making the code more concise
    Handling exceptions
    Opening files
X   Managing the resources 
"""

########################################################################################################################
"""
Context manager
Full name

Rewrite the code below using the with keyword.
Make sure to use the same variable names for file objects!
f1 = open('name.txt')
f2 = open('surname.txt')
f3 = open('full_name.txt', 'w')

name = f1.read()
surname = f2.read()

full_name = name + ' ' + surname

f3.write(full_name)
"""
# with open('name.txt') as f1:
#     name = f1.read()
# with open('surname.txt') as f2:
#     surname = f2.read()
# with open('full_name.txt', 'w') as f3:
#     full_name = name + ' ' + surname
#     f3.write(full_name)

########################################################################################################################
"""
Context manager
10 files

Create 10 files, named file1.txt, file2.txt and so on till file10.txt. The files should contain the number 
corresponding to their name. So, file1.txt should contain one line with number 1, file2.txt — one line with number 2 
and so on.
"""
# for i in range(1, 11):
#     with open(f"file{i}.txt", 'w') as f:
#         f.write(f"{i}")
#     f.close()

########################################################################################################################
"""
Context manager
Yearly income

The file salary.txt contains a monthly salary of all employees in the company. Assume that the monthly salary is fixed 
and each employee is mentioned once. So, each number (an integer) written on a separate line corresponds to a particular 
employee:

3500
4780
6666
...

According to the example, the first worker gets 3,500 per month, the second one 4,780 and the third 6,666, etc.

Calculate how much each employee earns per year and save their yearly income to a file salary_year.txt. Similarly to the 
original file, each income should be on a separate line. Preserve the order as it helps identify an employee.
"""
# result = 0
# year_in_months = 12
# with open('salary.txt', 'r') as f1:
#     with open('salary_year.txt', 'w') as f2:
#         for number in f1:
#             result = int(number.replace("\n", "")) * year_in_months
#             f2.write(str(result)+"\n")

########################################################################################################################
"""
Context manager
Equivalent

Which option below is equivalent to the following code?
with open('alien.txt', 'w') as f:
    f.write('Take us to your leader!')
"""

########################################################################################################################
"""
Working with CSV
Useful PEP

What PEP is dedicated to CSV files?
    PEP 306
    PEP 20
    PEP 8
X   PEP 305
"""

########################################################################################################################
"""
Working with CSV
Data entry

What is a data entry?
    It is a name of the library for working with CSV files in Python.
    It is a table in a CSV file.
X   In a CSV file, it is a simple line with some text information separated by commas.
"""

########################################################################################################################
"""
Working with CSV
The first line

What does the first line of a CSV file contain?
X   The titles of columns
    The separators
    The information for each column
    Meta-information about the file
"""

########################################################################################################################
"""
Working with CSV
What is CSV?

What does CSV stand for?
    comma-separated valuations
    comma-semicolon values
    column-separated values
X   comma-separated values
"""

########################################################################################################################
"""
Working with CSV
DictReader and DictWriter

    What is the main purpose of DictReader and DictWriter?
    They allow you to delete csv.
    They allow you to import the csv library.
    They allow you to read and write any type of files.
X   They allow you to work with information from csv-files as dictionaries.
"""

########################################################################################################################
"""
Working with CSV
Separator

What separator do programmers usually use in CSV files?
,
"""

########################################################################################################################
"""
Working with CSV
Finances

Now try to create a program for working with CSV files on your computer. We have attached a file where you can find the 
information on the financial statistics of a government in June 2019. Your task is to open this file, process the data, 
and count how many data entries have the FINAL status (it is specified in the STATUS column). Print the number in the 
box below.
"""
# # Repons à l'exercice
# with open("government-finance-statistics-general-government-year-ended-june-2019-csv.csv", "r") as f:
#     file_reader = csv.DictReader(f, delimiter=",")
#     counter = 0
#     for line in file_reader:
#         if line['STATUS'] == 'FINAL':
#             counter += 1
#     print(counter)


# In addition:

# Print all the line with header, value as a dictionary
# with open("government-finance-statistics-general-government-year-ended-june-2019-csv.csv", "r") as f:
#     file_reader = csv.DictReader(f, delimiter=",")
#     for line in file_reader:
#         print(line)
#         print(line['Series_reference'], line['Period'], line['Data_value'], line['STATUS'], line['UNITS'],
#               line['MAGNTUDE'], line['Subject'], line['Group'], line['Series_title_1'], line['Series_title_2'],
#               line['Series_title_3'], line['Series_title_4'], line['Series_title_5'])


# Print all the lines including the header line
# with open("government-finance-statistics-general-government-year-ended-june-2019-csv.csv", "r") as f:
#     file_reader = csv.reader(f)
#     for line in file_reader:
#         print(line)

########################################################################################################################
"""
Working with CSV
Spirits

We all know that too much drinking is bad. Scientists often create tables in which they try to fix the alcohol available
for people to track its further consumption. We have attached an example of this file: the availability of alcohol in 
December 2019. Your task is to create a program for reading this CSV file, then count the number of data records that 
have 0 in the MAGNITUDE column. Finally, print this number in the box below.
"""
# with open('alcohol-available-for-consumption-year-ended-december-2019-csv.csv', 'r') as f:
#     file_reader = csv.DictReader(f, delimiter=',')
#     counter = 0
#     for line in file_reader:
#         print(line)
#         if line['MAGNITUDE'] == '0':
#             counter += 1
#     print(counter)

########################################################################################################################
"""
Working with CSV
Easy code

Order the lines of the code that help you to read and print information from the given CSV file.
students = open('students.csv', 'r', encoding='utf-8')
for line in students:
  separ_line = line.split(',')
  print(separ_line)
students.close()
"""

########################################################################################################################
"""
Working with CSV
Cities

Imagine you have a file where you can find information about the population and the sizes of all the cities in the 
world. City names are in the first column, their population is in the second one, the sizes are in the third one. You 
also have a code that computes the number of cities that have a population of more than 1000000 people. Unfortunately, 
the lines got mixed. You need to put them in the right order.
"""
# with open("cities", "r", encoding='utf-8') as cities:
#     count = 0
#     for line in cities:
#         split_line = line.split(",")
#         if int(split_line[1]) > 1000000:
#             count += 1

########################################################################################################################
"""
List
From X to a list

Select all Python objects that can be converted to a list with the list() function.
    Floating-point numbers
X   Strings
    Integer numbers
"""

########################################################################################################################
"""
List
List element

Match the list from the first column with the correct number of its elements.
[]                                  0
["brouhaha"]                        1
list("canoodle")                    8
[1, [2, 3], [4, 5], 6]              4
["1", "2", "3", "5", "8", "13"]     6
"""

########################################################################################################################
"""
List
How to create a list

Select all ways to create a list in Python.
    Via single or double quotation marks: '', "".
X   Via square brackets: [].
    Via curly brackets: {}.
x   Via list() function.

"""

########################################################################################################################
"""
List
Reveal the hidden

There's a variable hidden, which contains a list. Count the number of elements in this list and print the length of the 
list.    
"""
# # the following line reads the input and converts it into a list; do not modify it, please
# hidden = list(input())
# print(len(hidden))

########################################################################################################################
"""
List
What does len() do?
We applied the function len() to a list. What does the returned number stand for?

    The index of the last element in the list.
    The last element appended to the list.
X   How many elements there are in the list.
    How many unique elements there are in the list.
    How many different data types there are in the list.
"""

########################################################################################################################
"""
List
An empty list!
Write a program that will create an em_list variable. Assign an empty list to it, then print your empty list.
"""
# em_list = []
# print(em_list)

########################################################################################################################
"""
List
Find the right list
What is the correct representation of a list?

    [5, 6, 7, "8"}
X   [1, '2', 3, "4"]
    [1 2 3 4 5]
    ('1', 2, 3, '4')
"""

########################################################################################################################
"""
List
Single-element list
Helen wants to create a list containing only one element – a string with her name. To do this, she wrote this line of code:
name = list('Helen')
However, when she prints the list, her name appears broken into several one-letter strings:
print(name)
# ['H', 'e', 'l', 'e', 'n']
Why? Fix this code.
"""
# name = list(['Helen'])
# print(name)

########################################################################################################################
"""
List
Recall len()
Select all datatypes the len() function can be used with.

X   a list
    an integer
    a Boolean
X   a string
    a float
"""

########################################################################################################################
"""
List
Features of a list
Select false statement(s) about lists:

    lists can store different types of elements (e.g. strings and numbers)
    the elements' order matters
    lists can store duplicate values
X   lists cannot be empty
"""

########################################################################################################################
"""
List
Creating a list

    ()
X   list()
    list[123]
X   ["1", "2", "3"]
X   [7, "wonders"]
"""

########################################################################################################################
"""
List
List from string
Write a program that constructs a list from the given string's symbols and then prints it.
So, if input_str, for example, contains a string "python", your program should turn it to the list 
['p', 'y', 't', 'h', 'o', 'n'] and print it.
"""
# input_str = input()
# print(list(input_str))

########################################################################################################################
"""
List
Syntax for the list

How to create a list containing numbers?
    (1, 2, 3, 4, 5)
    {1, 2, 3, 4, 5}
    list(1, 2, 3, 4, 5)
    "1 2 3 4 5"
X   [1, 2, 3, 4, 5]
"""

########################################################################################################################
"""
Operation with list
Merging lists

Write a function that would take two lists and return one merged list, e.g. for the lists 
['Washington, D.C.', 'Chicago', 'New York'] and ['Los Angeles', 'Las Vegas'] 
the result should be ['Washington, D.C.', 'Chicago', 'New York', 'Los Angeles', 'Las Vegas'].

You can use any of the methods described in the theory. Don't print the merged list, just return it at the end of the 
function.
"""
# list_one = ['Washington, D.C.', 'Chicago', 'New York']
# list_two = ['Los Angeles', 'Las Vegas']
# def merge_lists(list_one, list_two):
#     return list_one + list_two
# print(merge_lists(list_one, list_two))

########################################################################################################################
"""
Operation with list
The sequence of operations

Given a list with several numbers:
numbers = [2, 2, 4, 1, 1, 3, 5]
Here is a sequence of operations performed on this list:
numbers.remove(1)
numbers.extend([0])
numbers.append(len(numbers))
numbers.remove(5)
numbers.append(5)
What does the result list look like?

    [5, 7, 0, 2, 4, 1, 3, 5]
    [2, 2, 4, 3, 0, 6, 5]
    [[0], 2, 2, 4, 1, 3, 7, 5]
X   [2, 2, 4, 1, 3, 0, 7, 5]
"""

########################################################################################################################
"""
Operation with list
From one list to another

Which method can add all elements from one list to another within a single operation?
    append()
X   extend()
    insert()
    add()
"""

########################################################################################################################
"""
Operation with list
Appending and inserting elements 

What will be the output of the following code?
letters = ['a', 'b']
letters.append('c')
letters.insert(1, 'd')
letters.insert(2, 'e')
letters.append('f')
print(letters)

    ['d', 'e', 'c', 'f']
X   ['a', 'd', 'e', 'b', 'c', 'f']
    ['a', 'b', 'c', 'd', 'e', 'f']
    ['d', 'a', 'e', 'b', 'c', 'f']
    ['d', 'e', 'a', 'b', 'c', 'f']

"""

########################################################################################################################
"""
Operation with list
Adding and removing elements

What will be the output of the following code?
my_list = [0, 1, 2, 1, 2]
my_list.append(0)
my_list.remove(2)
my_list.append(1)
my_list.remove(0)
print(my_list)

    [0, 1, 2, 1, 1]
    [0, 1, 1, 0, 2]
X   [1, 1, 2, 0, 1]
    [1, 1, 0, 2, 1]
"""

########################################################################################################################
"""
Operation with list
List membership

Write down the name of the variable (type_1 or type_2) that should be placed instead of three dots to result in True 
value of the membership test.
container_types = ['list', 'set', 'dict', 'string']
type_1 = 'bool'
type_2 = 'list'
print(... in container_types)
>> type_2
"""
# container_types = ['list', 'set', 'dict', 'string']
# type_1 = 'bool'
# type_2 = 'list'
# print(type_2 in container_types)

########################################################################################################################
"""
Operation with list
Shopping list

She wrote down things that she would like to buy at the grocery store. All of these things are in the given 
shopping_list variable, it has already been created.

When sneaking a look at the fridge, she noticed that there was some milk left and changed her mind about buying a new 
one. It got crossed out of the list. And, finally, she added a "brownie", and it became the last item Ruth had put on 
the list.

Now you try to remove "milk" from the given variable shopping_list, and add "brownie" to it.

DON'T print the results.
"""
# shopping_list = []
# shopping_list.remove("milk")
# shopping_list.append("brownie")

########################################################################################################################
"""
Operation with list
Delete it!

Let's say there's a list values = ['delete me', 'do not', 'delete me'].
What are the correct ways to delete the first element?

X   values.pop('delete me')
    values.del(0)
    values.remove('delete me')
    values.pop()
X   del values[0]
    values.remove(0)

"""

########################################################################################################################
"""
Operation with list
Matching results

There's a list with student's grades during a semester: grades = [8, 5, 7, 9, 8, 6, 5, 7, 6, 7, 8, 8, 7, 9]. Compute 
the expressions on the left to their results on the right.

10 in grades                False
grades.index(8, 6, 10)      ValueError
grades.index(7, 5)          7
grades.count(5)             2
"""
# grades = [8, 5, 7, 9, 8, 6, 5, 7, 6, 7, 8, 8, 7, 9]
# print(10 in grades)
# # print(grades.index(8, 6, 10))
# print(grades.index(7, 5))
# print(grades.count(5))
########################################################################################################################

"""
List Index
Error indexes

What lines are correct and do not throw an error?
Remember that the first square brackets denote a list, and all the rest include indexes.
X   [1, 2, 3, 4][-4]
    [1, 2, 3][-4]
X   [1, 2, 3, 4][3]
    [1, 2, 3, 4][4]
"""

########################################################################################################################
"""
List Index
Different alphabets

We have created a variable alphabet for lowercase alphabets of different languages. Your task is to print the 15th 
letter of this string.
"""
# # don't modify the variable below, please
# alphabet = input()
# # put your python code here
# print(alphabet[14])

########################################################################################################################
"""
List Index
Guess the outcome

Guess what happens when we run the code below:
boat = [1, 2, 3, 'dog']
print(boat[4])

    boat[4]
    1
    'dog'
X   IndexError
"""

########################################################################################################################
"""
List Index
The syntax

What is the syntax for accessing an element of a list by an index?
    index(list)
X   list[index]
    list(index)
    list.index()
    list.element(index)
"""

########################################################################################################################
"""
List Index
Initials

Find the first letter of a person's name and print it out.
Make use of the variable name that stores a string.
"""
# # the following line reads the name from the input, do not modify it, please
# name = input()
#
# # your code here
# print(name[0])

########################################################################################################################
"""
List Index
Find the third number

The list prices contains several numbers. Print the third number from this list.
For example, if prices looks like [170, 309, 224, 991, 4000], you should print 224.
"""
# # the following line reads the list from the input, do not modify it, please
# # split() helps to convert your input into a list
# prices = input().split()

# # please work with the list 'prices' here
# print(prices[2])

########################################################################################################################
"""
List Index
Calculator

A calculator stores the result of the previous calculations in its memory as a list. You want to access the last 
calculation, but you don't know how many calculations there were in total.

Work with the calculations variable and print the last calculation.
"""
# # please, don't modify this variable
# calculations = input().split()
#
# # complete the code
# print(calculations[-1])

########################################################################################################################
"""
List Index
Tail

Sentences generally end with a certain punctuation mark: a period ., an exclamation point !, or a question mark ?.
Find out which of these symbols marks the end of a string stored in the variable sentence and print it out.
"""
# sentence = input()
#
# # now print the last symbol of `sentence`
# print(sentence[-1])

########################################################################################################################
"""
List Index
Find an error

Which line of code will cause an error?
fruits = ["apple", "pear", "orange", "mango", "peach"]

print(fruits[0])     # 3
print(fruits[-5])    # 4
print(fruits[5])     # 5
print(fruits[4-2])   # 6

    3
    4
X   5
    6
"""

########################################################################################################################
"""
List Index
Sequence items

Given the sequence seq, match its elements with corresponding indexes.
the next-to-last element    len(seq) - 2
the last element            len(seq) - 1
the second element          -len(seq) + 1
the first element           -len(seq)
"""

########################################################################################################################
"""
List Index
Modifying data

Lists, unlike strings, are mutable. We can use that to modify their data with indexes.
There is a list planets with the names of the Solar system planets. However, instead of the 5th planet, there's an X. 
Replace it with the real name of the 5th planet. Do not forget, the first letter must be a capital one!
Note that the list planets has already been defined, you just need to change one element. Mind its index!
"""
# # change the name of the 5th planet in planets
# planets[4] = "Jupiter"

########################################################################################################################
"""
List Index
Lucky tickets are a kind of mathematical entertainment. A ticket is considered lucky if the sum of the first 3 digits 
equals the sum of the last 3 digits of the ticket number.

You are supposed to write a program that checks whether the two sums are equal. The code snippet below displays "Lucky" 
if they are and "Ordinary" if they are not.

However, some parts of the code are missing. Fill in the blanks to make it work!

Input: a string of 6 digits.

Output: either "Lucky" or "Ordinary" (without quotes).

Make sure that you are NOT concatenating strings. To do this, convert each digit in the ticket number to an integer. 
Don't forget to use proper indices.
"""
# # Save the input in this variable
# ticket = [int(item) for item in list(input())]
#
# # Add up the digits for each half
# # half1 = ticket[0] + ticket[1] + ticket[2]
# half1 = sum(ticket[0:3])
#
# # half2 = ticket[-1] + ticket[-2] + ticket[-3]
# half2 = sum(ticket[:-4:-1]) # on parcours en sens inverse la liste(-1) et on va du dernier element à l'element -3 il
#                             # faut mettre -4 car dans la borne sup. est toujours exclu
#
# print(half1, half2)
#
# # Thanks to you, this code will work
# if half1 == half2:
#     print("Lucky")
# else:
#     print("Ordinary")

########################################################################################################################
"""
Nested lists
A lot of nested lists

Let's write a program that will read a positive integer n from the input and create a nested list containing the inner 
list [1, 2] repeated n times.
"""
# n = int(input())
# my_list = [[1, 2]]
# print(my_list*n)

########################################################################################################################
"""
Nested lists
Accessing elements of a matrix

Let's say we have a matrix M:
M = [[34, 77, 8,  45],
     [10, 15, 93, 57],
     [78, 82, 19, 98]]
     
Write a code that will print the element in the first column and the third row. Mind that the size and the number of 
elements in the test matrix will be the same, but the integers in it will be different.    
"""
# the variable M is already defined
# M = [[34, 77, 8,  45],
#      [10, 15, 93, 57],
#      [78, 82, 19, 98]]
#
# print(M[2][0])

########################################################################################################################
"""
Nested lists
Double for
Which of the following list comprehension will do the same as the code below?

country_list = [["Moscow", "Cheboksary", "Sochi"], ["Paris", "Lyon", "Nice"],
                ["New York", "Dallas", "San Francisco"]]

long_cities = []
for country in country_list:
    for city in country:
        if len(city) >= 6:
            long_cities.append(city)
            
long_cities = [city if len(city) >= 6 for city in country for country in country_list]
long_cities = [city for city in country for country in country_list if len(city) >= 6]
long_cities = [city for country in country_list for city in country if len(city) >= 6]
long_cities = [city if len(city) >= 6 for country in country_list for city in country]
"""
# country_list = [["Moscow", "Cheboksary", "Sochi"], ["Paris", "Lyon", "Nice"],
#                 ["New York", "Dallas", "San Francisco"]]
#
# long_cities = [city for country in country_list for city in country if len(city) >= 6]

########################################################################################################################
"""
Nested lists
The length of a list
Enter the length of the given list.

numbers = [3, 4, [5, 1], 1, [7, 8]]
>> 5
"""

########################################################################################################################
"""
Nested lists
Length of a list
What is the length of the list below?

my_list = [1, [2, [3, 4]]] 
print(len(my_list))
>> 2
"""
# my_list = [1, [2, [3, 4]]]
# print(len(my_list))

########################################################################################################################
"""
Nested lists
Choosing elements from nested lists
Imagine that the list below shows a part of a cafe's menu where the first element in each nested list is the name of a 
dish, the second element is the number of people the dish is aimed at, and the third number is its cost in the local 
currency.

menu = [["pizza", 4, 20], ["soup", 1, 8], ["ice cream", 2, 4], ["toasts", 2, 10]]
Let's say a customer wants to select what to order using Python, and they have written the following code:
what_to_order = [dish[0] for dish in menu if dish[1] >= 2 and dish[2] < 15]
print(what_to_order)
Which output will this expression yield?
X   ["ice cream", "toasts"]
    [["ice cream"], ["toasts"]]
    [["ice cream", 2, 4], ["toasts", 2, 10]]
    "ice cream", "toasts"
"""

########################################################################################################################
"""
Nested lists
A very nested list
Write a program that takes three strings as input and then constructs and prints a nested list from them – with first 
string as a first element, a list containing only second string as a second element and a list containing another list 
containing a third string as a third element.
"""
# str_1 = input()
# str_2 = input()
# str_3 = input()
# result = [str_1, [str_2], [[str_3]]]
# print(result)

########################################################################################################################
"""
Nested lists
Fill the blanks
Below you can see the code that chooses some elements from one list and appends them to another:
for a in x:
    for el in a:
        if el > 0:
            els.append(el)            
Fill in the blanks in the code below so that list comprehension produces the same result as the code above.
# here we create the initial list from the input, please do not modify this line
x = json.loads(input())
els = [el for a in x for el in a if el > 0]
"""

########################################################################################################################
"""
Nested lists
Problem of the day
Conditions & nested lists
"""
# students = [["Jane", "B"], ["Kate", "B"], ["Alex", "C"], ["Elsa", "A"], ["Max", "B"], ["Chris", "A"]]
# students = [a[0] for a in students if a[1] == "A"]
# print(students)

########################################################################################################################
"""
List Slicing
Theory
"""
# text = 'hello'
# print(text[0:999])
#
# snakes = ['python', 'cobra', 'viper']
# print(snakes[2][2:])
"""
>> hello
>> per
"""

########################################################################################################################
"""
List Slicing
You've got mail!

Email addresses usually have the same structure: someone@somewhere.com (or any other domain).
You are working with a database of email addresses from an imaginary email service yougotmail.com. Create a program 
that would separate the local-part that precedes the @ sign from the rest of the address.
The input format:
The email address. It is stored in the variable email, you do not need to work directly with input.
The output format:
The local-part of the address.
"""
# email = 'someone@yougotmail.com'
# result = email.split('@')
# print(result[0])
#
# email = 'john.andrew.smith@yougotmail.com'
# result = email.split('@')
# print(result[0])

########################################################################################################################
"""
List Slicing
How to reverse a string
Here is a string written in Python:

s = "wolf"

How can you reverse it using slicing?
    s[-1::]
    s[:-1]
    s[-1:1]
X   s[::-1]
"""

########################################################################################################################
"""
List Slicing
Email

Here is a string representing an email:

email = 'python@hyper.skill'
What is a result of this slicing operation?
email[3:-3:3]
X   'h@p.'
    'hnhprs'
    'hi'
    'thyrk'

"""

########################################################################################################################
"""
List Slicing
Slice length

What is the output of the following code?
len([8, 18, 28, 38, 48, 58, 68, 78, 88][-6:-2:3])
    0
    1
X   2
    3
"""
# print([8, 18, 28, 38, 48, 58, 68, 78, 88][-6:-2:3])
# print(len([8, 18, 28, 38, 48, 58, 68, 78, 88][-6:-2:3]))

########################################################################################################################
"""
List Slicing
All 5

We have created a list of integers called numbers. Use this list to create a new one in which all the integers are 
divisible by 5. Print it.
numbers = [5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20,
           21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32, 33, 34, 35, 
           36, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 50, 
           51, 52, 53, 54, 55, 56, 57, 58, 59, 60, 61, 62, 63, 64, 65, 
           66, 67, 68, 69, 70, 71, 72, 73, 74, 75, 76, 77, 78, 79, 80,
           81, 82, 83, 84, 85, 86, 87, 88, 89, 90, 91, 92, 93, 94, 95, 
           96, 97, 98, 99, 100]
"""
# numbers = [5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20,
#            21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32, 33, 34, 35,
#            36, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 50,
#            51, 52, 53, 54, 55, 56, 57, 58, 59, 60, 61, 62, 63, 64, 65,
#            66, 67, 68, 69, 70, 71, 72, 73, 74, 75, 76, 77, 78, 79, 80,
#            81, 82, 83, 84, 85, 86, 87, 88, 89, 90, 91, 92, 93, 94, 95,
#            96, 97, 98, 99, 100]
# new_list = numbers[::5]
# print(new_list)

########################################################################################################################
"""
List Slicing
Big index

Let's say we have a list series = ['Game of Thrones', 'Riverdale', 'Shameless', 'Doctor Who', 'Friends'].
What will be the result of the following expression: print(series[3:5])?

    ['Shameless', 'Doctor Who', 'Friends']
    There will be an error
    ['Doctor Who', 'Friends']
X   ['Doctor Who', 'Friends', '']
"""
# series = ['Game of Thrones', 'Riverdale', 'Shameless', 'Doctor Who', 'Friends']
# print(series[3:5])

########################################################################################################################
"""
List Slicing
Palindrome check

To find out if a word is a palindrome we would need to check if it reads the same forward and backward.
The condition for that check has already been written in the code below, but the parts that need to be compared haven't 
been defined yet. Finish the code by defining variables forward and backward.
The variable that stores the word in question is called word. 
"""
# word = input()
# forward = word[:]
# backward = word[::-1]
# print(forward, backward)
# if forward == backward:
#     print("Yes")
# else:
#     print("No")

########################################################################################################################
"""
List Slicing
Default values

Here is the full syntax for slicing:
sequence[start:stop:step]
What are the default values for start, stop and step?

X   start: 0
start: 1
X   stop: len(sequence)
stop: len(sequence) - 1
step: 0
X   step: 1
"""

########################################################################################################################
"""
List Slicing
Understanding different forms

Match different forms of slicing with their descriptions.
sequence[start:]    makes the full copy of the sequence
sequence[:end]      takes elements from start to end-1
sequence[::step]    takes all elements with a given step
sequence[:]         takes elements from start to the last element

Solution:
sequence[start:]    takes elements from start to the last element
sequence[:end]      takes elements from start to end-1
sequence[::step]    takes all elements with a given step
sequence[:]         makes the full copy of the sequence
"""

########################################################################################################################
"""
List Slicing
Quotation marks

Given a string in quotes, print it out without quotation marks.
 Sample Input 1:
"python"

Sample Output 1:
python
"""
# # work with this variable
# string = input()
# print(string[1:-1])

########################################################################################################################
"""
Sorting a List
Which method did it?

Select the operation(s) below that can turn this list:
numbers = [1, 6, -9, 2.5, 4, 0]
into this:
[6, 4, 2.5, 1, 0, -9]
within a single operation.

    sorted_numbers = sorted(numbers)
    numbers.sort(reverse=False)
    sorted_numbers = reversed(numbers)
X   numbers.sort(reverse=True)
X   sorted_numbers = sorted(numbers, reverse=True)
    numbers.reverse()
"""

########################################################################################################################
"""
Sorting a List
Can’t sort my list!

Nathan has a list that contains the names of all his friends. He wants to sort it in alphabetical order and then print 
it. However, when he executes this code:
my_list = ['Mary', 'Bob', 'Pete', 'Jane']
my_list = my_list.sort()
print(my_list)
only None is printed! Help him – find a broken line in his code, fix it and write down the corrected line.
my_list.sort()
"""
# my_list = ['Mary', 'Bob', 'Pete', 'Jane']
# my_list.sort()
# print(my_list)

########################################################################################################################
"""
Sorting a List
Gaming task

Little Niki played with toys and put them into a list to keep track of them.
toys = ['Marinet', 'Toby', 'Kitty']
Then, she decided to add another toy, named "Chubby", after which she sorted the list. As a result, she got the 
following:
['Toby', 'Kitty', 'Chubby', 'Marinet']
Unfortunately, Niki forgot to save the code. Restore it to get the same result!
The input you will get is a list of four toy names, including "Chubby". You should sort the list and print it. Pay 
attention to the sorting function!
"""
# # the following line reads the list from the input, do not modify it, please
# toys = input().split()
#
# # your code below
# toys.sort(key=len)
# print(toys)

########################################################################################################################
"""
Sorting a List
Unsortable!

Select a list that can’t be sorted via sort() method.
    [1, 2, 3, 4, 5, 6]
    [1, 6, -9, 2.5, 4]
    ['!!!', "Aaa", "Aab", 'foo', 'bar', "foobar"]
X   [5, 1, -2, "45", 0]
"""

########################################################################################################################
"""
Sorting a List
Sort the flowers

Mrs. Smith works in a flower shop. Here you can see a table with the flower prices:
flower 	price, $
rose 	0.75
lily 	1.3
tulip 	1
hyacinth 	0.70

Mrs. Smith needs to sort prices in the descending order. Write a program to help her!
The input you will get is a list of numbers, e.g. [0.75, 1.3, 1, 0.70]. The result should be this list sorted in the 
descending order, [1.3, 1, 0.75, 0.7] for the example above. Please print the sorted list in the end.
"""
# # the following line creates a list from the input, do not modify it, please
# prices = [float(price) for price in input().split()]
# # your code below
# prices.sort(reverse=True)
# print(prices)

########################################################################################################################
"""
Sorting a List
Sort numbers by the remainder

You are given a list of numbers nums. Sort the list by the remainder of dividing by three and print the result. Use a 
lambda function as key if possible.
"""
# nums = [int(num) for num in list(input())]
#
# # write your code here
# nums.sort(key=lambda x: x % 3)
# print(nums)

########################################################################################################################
"""
Sorting a List
Sort the numbers

Below you can see a list of strings called numbers. Sort it in the descending order as strings (alphabetically) and 
print the resulting list.
"""
numbers = ["77", "145", "987", "2095", "6", "371", "4999", "81"]

# # sort numbers
# numbers.sort(reverse=True)
# print(numbers)

########################################################################################################################
"""
Sorting a list
Pass the password

Tim chooses a password for social networks so that it be difficult to crack. He wrote down the options in a list:

passwords = ['0vbno0re', 'ad12', 'fgghut', '4qp', 'qwerty']

Now, he wants to determine the level of their security depending on the length. A password that has a length of 1 is 
the most unsafe one, while a password of length 10 is considered the best one.

Given a list of passwords, sort them from the least safe to the safest one. As the output, print all passwords together 
with their length. For the example above, you should print:

4qp 3
ad12 4
fgghut 6
qwerty 6
0vbno0re 8
"""
# # the follwing line reads the list from the input, do not modify it, please
# passwords = input().split()
#
# # your code below
# passwords.sort(key=len)
# for password in passwords:
#     print(password, len(password))
#
# # Solution 2
# for password in sorted(passwords, key=len):
#     print(password, len(password))

########################################################################################################################
"""
Sorting a list
Changed beyond recognition

Dany created the following list:

dragons = ['Rudy', 'Targent', 'Aggie']

and then played with it for a while. She performed some operations on it, so the result looks like this:

['Targent', 'Aggie', 'Rudy']

Now Dany is confused and wants to understand what exactly she has done. Help her out! Place the operations in the 
correct order – the order in which they were used to get ['Targent', 'Aggie', 'Rudy'] from ['Rudy', 'Targent', 'Aggie'].
dragons = sorted(dragons, key=len)
dragons.reverse()
dragons.sort()

Solution
dragons.sort()
dragons = sorted(dragons, key=len)
dragons.reverse()
"""

########################################################################################################################
"""
Sorting a list
Maximum and minimum

Write a program that receives three integers as input and prints them in the following order: the maximum, the minimum, 
and then the remaining number, all in one line. Note that there can be duplicate numbers in the input.
"""
# # the following code creates a list from input, please do not modify it
# ints = [int(num) for num in input().split()]
#
# # your solution here
# max_list = sorted(ints, reverse=True)
# print(max_list[0], max_list[2], max_list[1])

########################################################################################################################
"""
List comprehension
Else comprehension
# the following line reads the list from the input, do not modify it, please
old_list = [int(num) for num in input().split()]

binary_list = # your code here
print(binary_list)
"""
# old_list = [int(num) for num in input().split()]
# print(old_list)
#
# binary_list = [1 if i > 0 else 0 for i in old_list]
# print(binary_list)

########################################################################################################################
"""
List comprehension
Plus one
You are given a list of strings containing integer numbers. Print the list of their values increased by 1.
E.g. if list_of_strings = ["36", "45", "99"], your program should print the list [37, 46, 100].
The variable list_of_strings is already defined, you don't need to work with the input or define it again.
"""
# list_of_strings = [int(num) for num in input().split()]
# print([int(i) + 1 for i in list_of_strings])

########################################################################################################################
"""
List comprehension
Length
Given a list of words in the code below, create a list of lengths of those words and print it.
words = ["apple", "it", "creek", "pelican", "subsequent", "horse", "apothecary"]
"""
# words = ["apple", "it", "creek", "pelican", "subsequent", "horse",
#          "apothecary"]
# list_of_length = [len(word) for word in words]
# print(list_of_length)

########################################################################################################################
"""
List comprehension
Boundary
Write a program that divides numbers into two lists depending on whether they are greater than or less than 5. You don't
 have to include number 5 itself.
A sequence of numbers has been read from the input for you.
# work with a list from this variable
numbers = [int(n) for n in input()]

# change the next two lines
less_than_5 = ...
greater_than_5 = ...

# printing your results
print(less_than_5)
print(greater_than_5)
"""
# # work with a list from this variable
# numbers = [int(n) for n in input().split()]
#
# # change the next two lines
# less_than_5 = [number for number in numbers if number < 5]
# greater_than_5 = [number for number in numbers if number > 5]
#
# # printing your results
# print(less_than_5)
# print(greater_than_5)

########################################################################################################################
"""
List comprehension
Even numbers

Write a program that takes a list of numbers, creates another list of even numbers from the first list, and prints it.

E.g. if my_numbers = [1, 2, 3, 4, 5], then your program should print the list [2, 4].

# the following line reads the list from the input; do not modify it, please
my_numbers = [int(x) for x in input().split(" ")]

# work with the variable 'my_numbers'
"""
# # the following line reads the list from the input; do not modify it, please
# my_numbers = [int(x) for x in input().split(" ")]
#
# # work with the variable 'my_numbers'
# print([number for number in my_numbers if number % 2 == 0])


########################################################################################################################
"""
List comprehension
The mean

Given a numeric sequence at the input, create a list in which each number will be a digit from this input string. Then
use this list to calculate the arithmetic mean, that is, the sum of all the digits divided by their total count.
Don't forget to print your result.
"""
# test = [float(i) for i in list(input())]
# print(sum(test) / len(test))


########################################################################################################################
"""
List comprehension
How many days?

The list seconds is a list of numbers that represent seconds. Convert the number of seconds to full days and print the 
list that contains these values. The number of full days should be an int value.
seconds = [86400, 1028397, 8372891, 219983, 865779330, 3276993204380912]
# create a list of days here
"""
# seconds = [86400, 1028397, 8372891, 219983, 865779330, 3276993204380912]
# nb_days = [second // (3600 * 24) for second in seconds]
# print(nb_days)

########################################################################################################################
"""
List comprehension
Threefold

Print a list of numbers from 1 to 1000 that can be divided by 3.
"""
# list_of_numbers = [print(x) for x in range(0, 1001) if x % 3 == 0]
#
# # or to get it validated
# list_of_numbers = [x for x in range(1, 1001) if x % 3 == 0]
# print(list_of_numbers)

########################################################################################################################
"""
List comprehension
Vowels

Read a string from the input and print a list of vowels that belong to that string.
All vowels are enlisted in a variable of the same name.

vowels = 'aeiou'
# create your list here
"""
# vowels = 'aeiou'
# create your list here
# string = input()
# vowels_list = [letter for letter in string if letter in vowels]
# print(vowels_list)

########################################################################################################################
"""
List comprehension
Odd numbers

Read a string with digits from the input and convert each number to an integer. Create a list in which you should
include only odd digits. Finally, print what you'll get.

Sample Input 1:
87113420

Sample Output 1:
[7, 1, 1, 3]
"""
# list_numbers = input()
# print([int(i) for i in list_numbers if int(i) % 2 == 1])

########################################################################################################################
"""
List comprehension
A list of words

Write a program that takes a list with words, creates a new list of words that start with the letter "a" in the first
list, and prints it.
Some words may begin with the capital A! Leave them in their original form in the resulting list.
E.g. if words = ['apple', 'pear', 'banana', 'Ananas'], then your program should print the list ['apple', 'Ananas'].
The list with words is already defined: you can access it using the variable words.
"""
# words = ['apple', 'pear', 'banana', 'Ananas']
# words = [word for word in words if word.startswith("a") or word.startswith("A")]
# print(words)

########################################################################################################################
"""
List comprehension
Running total

Given a numeric sequence, create a list in which each number will be a digit from this input string. Then use this list
to calculate the running total, or cumulative sum. Essentially, it's a new list of partial sums that gets updated every
time a new element appears.
For example, we can transform the list [1, 2, 3] to [1, 1 + 2, 1 + 2 + 3], which equals to [1, 3, 6].
"""
# Solution 1
# prior python 3.8


# list_numbers = list(input())
# list_beta = [int(number) for number in list_numbers]
# j = 0
# result=[]
# for i in list_beta:
#     i += j
#     j = i
#     result.append(i)
#
# print(result)


# Solution 2
# works only for python 3.8 and above
# list_numbers = list(input())
# result = 0
# print([result := result + int(list_numbers[i]) for i in range(len(list_numbers))])

# Solution 3
# numbers = [int(x) for x in list(input())]
# new_list = [sum(numbers[:i + 1]) for i, x in enumerate(numbers)]
# print(new_list)

# Rappel
# test = [0, 1, 2, 3, 4, 5]
# print(test[:6])

"""
List comprehension
Else comprehension

Write a program that converts a given list into a list with binary values: either 1 or 0. If the number in the initial 
list was greater than 0, in the binary list there should be 1, and if the number was less or equal, in the new list you 
should write 0. Naturally, use the list comprehension and print the result.
Given list: a list with integer numbers, e.g. [5, 0, 4, -10].
Output list: a list consisting of ones and zeros, e.g. [1, 0, 1, 0].
"""
# # the following line reads the list from the input, do not modify it, please
# old_list = [int(num) for num in input().split()]
#
# binary_list = [1 if num > 0 else 0 for num in old_list]
# print(binary_list)

# >> 5 0 4 -10
# >> [1, 0, 1, 0]
########################################################################################################################
"""
List comprehension
Plus one

You are given a list of strings containing integer numbers. Print the list of their values increased by 1.
E.g. if list_of_strings = ["36", "45", "99"], your program should print the list [37, 46, 100].
The variable list_of_strings is already defined, you don't need to work with the input or define it again.
"""
# # please work with the variable 'list_of_strings'
# list_of_strings = [10, 100, 1000, 10000]
# print([int(result)+1 for result in list_of_strings])

########################################################################################################################
"""
List comprehension
Syntax with a condition

Match a list comprehension expression with the value of new_list where the old_list is:
old_list = [1, 1, 2, 3, 5, 8, 13, 21, 34, 55]       
new_list = [x for x in old_list if x % 2 == 0]      [2, 8, 34]
new_list = [x for x in old_list if x >= 8]          [8, 13, 21, 34, 55]
new_list = [x*2 for x in old_list if x < 13]        [2, 2, 4, 6, 10, 16]
new_list = [x*2 for x in old_list if x % 2 == 1]    [2, 2, 6, 10, 26, 42, 110]
new_list = [x*2 for x in old_list]                  [2, 2, 4, 6, 10, 16, 26, 42, 68, 110]
"""

########################################################################################################################
"""
List comprehension
Length

Given a list of words in the code below, create a list of lengths of those words and print it.
"""

# words = ["apple", "it", "creek", "pelican", "subsequent", "horse",
#          "apothecary"]
# print([len(word) for word in words])

########################################################################################################################
"""
Dictionary
Count the marks

What will be printed out as a result of these actions?
new_dict = {"a": 6, "b": 3}
new_dict['a'] = new_dict['a'] / new_dict['b']
print(new_dict['a'] + new_dict['b'])

    5
X   5.0
    9.0
    The error will occur
    6
"""

########################################################################################################################
"""
Dictionary
Create

Let's say you asked your friends to name their favorite flowers: now you know that Alex likes 'field flowers', Kate 
prefers 'daffodil', Eva adores 'artichoke flower', and Daniel loves 'tulip'.
Create a dict with the names as keys and flowers as values and print it.
"""
# dict = {'Alex': 'field flowers', 'Kate': 'daffodil', 'Eva': 'artichoke flower', 'Daniel': 'tulip'}
# print(dict)

########################################################################################################################
"""
Dictionary
Grocery list

Alice is going to the store and she's writing a list of what she should buy. First, she didn't want to buy many bananas 
but then she suddenly changed her mind. How can you help Alice and correctly update the value ofbananas in the following
shopping_list?

shopping_list = {'bananas': 5, 'oranges': 3, 'yogurt': 2, 'chicken breasts': 3, 'olive oil': 1}
shopping_list.bananas = 10
shopping_list['bananas'] = 10
bananas['shopping_list'] = 10
shopping_list[bananas] = 10

    shopping_list.bananas = 10
X   shopping_list['bananas'] = 10
    bananas['shopping_list'] = 10
    shopping_list[bananas] = 10
"""

########################################################################################################################
"""
Dictionary
Let's count

What will be the result of the code below:ple = {}
sample['a'] = 3
sample['b'] = 5
sample['c'] = -2
print(sample['a'] + sample['b'] + sample['c'])

>> 6
"""

########################################################################################################################
"""
Dictionary
Comprehension

Choose all the possible ways to create a dictionary.
Report a typo
Select one or more options from the list
empty_dict = []
empty_dict = dict{}
empty_dict = dict()
empty_dict = {}

    empty_dict = []
    empty_dict = dict{}
X   empty_dict = dict()
X   empty_dict = {}
"""

########################################################################################################################
"""
Dictionary
Nested

Imagine there are three children in a family and they all wrote down what they want to be when they grow up:
children = {'Emily': 'artist', 'Adam': 'astronaut', 'Nancy': 'programmer'}
Let's say you want to store not only the profession they have chosen but also their age: Emily is 5, Adam is 9, and 
Nancy is 14. To do so, you can create nested dictionaries for each key in the outer dictionary.
For each name, create a nested dictionary with the keys 'profession' and 'age', modify the dictionary children, 
but don't print it.
NB: write the age as an integer.
children = {'Emily': 'artist', 'Adam': 'astronaut', 'Nancy': 'programmer'}
"""
# # please work with the variable children
# children = {'Emily': 'artist', 'Adam': 'astronaut', 'Nancy': 'programmer'}
#
# # please work with the variable children
# children = {
#     'Emily': {
#         "profession": "artist",
#         "age": 5},
#     'Adam': {
#         "profession": "astronaut",
#         "age": 9},
#     'Nancy': {
#         "profession": "programmer",
#         "age": 14}}

########################################################################################################################
"""
Dictionary
Update info

Let's say you have a dict with children's names and professions they would like to have when they grow up:
children = {'Emily': 'artist', 'Adam': 'astronaut', 'Nancy': 'programmer'}
But what if someone's choice has changed? Say, Emily now wants to be a musician. Update the dict and print its new 
version.
"""
# please work with the variable children
# update the dict and print it
# children = {'Emily': 'artist', 'Adam': 'astronaut', 'Nancy': 'programmer'}
# children['Emily']='musician'
# print(children)

########################################################################################################################
"""
Dictionary
Adding an item

Let's say you have a dictionary matching your friends' names with their favorite flowers:
fav_flowers = {'Alex': 'field flowers', 'Kate': 'daffodil', 'Eva': 'artichoke flower', 'Daniel': 'tulip'}
Your new friend Alice likes orchid the most: add this info to the fav_flowers dict and print the dict.
NB: Do not redefine the dictionary itself, just add the new element to the existing one.
"""
# please work with the variable fav_flowers
# fav_flowers = {'Alex': 'field flowers', 'Kate': 'daffodil', 'Eva': 'artichoke flower', 'Daniel': 'tulip'}
# fav_flowers['Alice'] = 'orchid'
# print(fav_flowers)

########################################################################################################################
"""
Dictionary methods
Let's count

What will be the result of the code below?
sample = {}
sample['a'] = 3
sample['b'] = 5
print(sample['a'] + sample['b'] + sample.get('c', -2) + sample.get('a', 10))

>> 9
"""

########################################################################################################################
"""
Dictionary methods
Big family

It's a wedding day — two happy people are getting married.

We have a dictionary first_family that contains the names of the wife's family members. For example:

first_family = {"wife": "Janet", "wife's mother": "Katie", "wife's father": "George"}

And a similar dictionary second_family for the husband's family:

second_family = {"husband": "Leon", "husband's mother": "Eva", 
                 "husband's father": "Gaspard", "husband's sister": "Isabelle"}

Create a new dictionary that would contain information about the new family. Similar to the ones above, family members 
should be keys and their names should be values. First, update the new dictionary with elements from first_family and 
then from second_family. Don't forget to print it out.
"""

# new_family=dict()
# first_family = {"wife": "Janet", "wife's mother": "Katie", "wife's father": "George"}
# second_family = {"husband": "Leon", "husband's mother": "Eva", "husband's father": "Gaspard", "husband's sister": "Isabelle"}
#
# # Solution 1
# new_family.update(first_family)
# new_family.update(second_family)
# print(new_family)
#
# # Solution 2
# new_family_2 = {**first_family, **second_family}
# print(new_family_2)
########################################################################################################################
"""
Dictionary methods
All hands on deck

In a standard deck of cards, there are 13 of each suit. There are numbered cards (from 2 to 10) and face cards (Jack, 
Queen, King, and Ace). If we were to rank the face cards, Jack would be 11, Queen 12, King 13, and the Ace 14.

Write a program that calculates the average rank of one hand of cards. Don't forget to consider the rank of the face cards.

The input format:
Six values of cards, each on a separate line.

The output format:
The average rank of the hand.
"""
# # Solution 1
# cards_value = {"0": 0, "1": 1, "2": 2, "3": 3, "4": 4, "5": 5, "6": 6, "7": 7, "8": 8, "9": 9, "10": 10, "Jack": 11,
#                "Queen": 12, "King": 13, "Ace": 14}
# sum = 0
# counter = 0
# cards = input().split()
#
# for card in cards:
#     if card in ["0", "1", "2", "3", "4", "5", "6", "7", "8", "9", "10", "Jack", "Queen", "King", "Ace"]:
#         sum += cards_value[card]
#         counter += 1
#
# print(sum/counter)
#
# # Solution 2
# card_deck = {'2': 2, '3': 3, '4': 4, '5': 5, '6': 6, '7': 7, '8': 8, '9': 9, '10': 10,
#              'Jack': 11, 'Queen': 12, 'King': 13, 'Ace': 14}
#
# total = 0
# for _ in range(6):
#     card = input()
#     total += card_deck[card]
# print(total / 6)

########################################################################################################################
"""
Dictionary methods
The Double

Write a program that creates a dictionary, in which keys are latin letters, and values are their doubling:
{a: aa, b: bb, ..., z: zz}
Save this dictionary in the variable double_alphabet
"""
# import string
# double_alphabet = dict()
# for letter in string.ascii_lowercase:
#     double_alphabet[letter] = letter + letter

########################################################################################################################
"""
Dictionary methods
Last

We have the following code:

student = {'name': 'Kate', 'age': 20, 'specialty': 'biology'}
student.update(degree='bachelor')
print(student.get('age'))
print(student.pop('specialty'))
print(student.popitem())

This program will print several lines. What will the last one be?
student = {'name': 'Kate', 'age': 20, 'specialty': 'biology'}
student.update(degree='bachelor')
print(student.get('age'))
print(student.pop('specialty'))
print(student.popitem())

X   ('degree', 'bachelor')
    {'specialty': 'biology'}
    'biology'
    'bachelor'
"""

########################################################################################################################
"""
Dictionary methods
Cleaning

Tom is learning dictionaries in Python. Here are the dictionaries he has created:

first_dict = {'key1': 'value1', 'key2': 'value2'}
second_dict = first_dict
third_dict = second_dict

Below there's a table with several lines of code on the left. Match them to their output. 
	                    {}	        {'key1': 'value1', 'key2': 'value2'}

first_dict.clear()      x
print(third_dict)

second_dict = {}                        x
print(third_dict)
"""

########################################################################################################################
"""
Dictionary methods
Delete

What are the correct ways to delete an existing key from the dictionary? Choose one or several options.
 Select one or more options from the list 
del my_dict[key]
remove "key" from my_dict
my_dict.pop(my_dict["key"])
my_dict.delete("key")

X   del my_dict[key]
    my_dict.pop(my_dict["key"])
    remove "key" from my_dict
    my_dict.delete("key")
"""

########################################################################################################################
"""
Dictionary methods
Delete from squares

We have a dictionary squares which contains numbers as keys and their squares as values. It looks like this:

squares = {1: 1, 3: 9, 5: 25, ...}

Your task is to read the input as a key that needs to be deleted from the dictionary and print the value of this key. 
If the number read from the input does not exist in the dictionary, then you need to print There is no such key. 
Finally, print the squares dictionary on the next line after all the changes.

The dictionary squares is already defined for you.
"""
# # There is no such key
# squares = {1: 1, 3: 9, 5: 25, 6: 36, 8: 64, 10: 100, 11: 121, 15: 225}
# value = int(input())
# return_value = squares.pop(value, 'There is no such key')
# print(return_value)
# print(squares)

########################################################################################################################
"""
Dictionary methods
We Are What We Eat

Anthony keeps track of what he eats: he writes down how many calories are in his meals. Use the list of dictionaries to 
calculate the total amount of calories per day and print it.

The sample input will look like:
meals = [
        {"title": "Oatmeal pancakes with apple and cinnamon", "kcal": 370},
        {"title": "Italian salad with fusilli and ham", "kcal": 320},
        {"title": "Bulgur with vegetables", "kcal": 350},
        {"title": "Curd souffle with lingonberries and ginger", "kcal": 225},
        {"title": "Oatmeal with honey and peanuts", "kcal": 440}]
"""
# meals = [
#         {"title": "Oatmeal pancakes with apple and cinnamon", "kcal": 370},
#         {"title": "Italian salad with fusilli and ham", "kcal": 320},
#         {"title": "Bulgur with vegetables", "kcal": 350},
#         {"title": "Curd souffle with lingonberries and ginger", "kcal": 225},
#         {"title": "Oatmeal with honey and peanuts", "kcal": 440}]
#
# calories = 0
# for line in meals:
#     calories += line["kcal"]
# print(calories)

########################################################################################################################
"""
Operations with dictionary
Dictionary membership

Replace the dots with membership operators so that the output result would always be True.

sidekicks = {'Don Quixote': 'Sancho Panza', 'Batman': 'Robin', 'Holmes': 'Watson'}
print('Sancho Panza' ... sidekicks)
print('Batman' ... sidekicks)
print('Han Solo' ... sidekicks)
print('' ... sidekicks)

Solution
sidekicks = {'Don Quixote': 'Sancho Panza', 'Batman': 'Robin', 'Holmes': 'Watson'}
print('Sancho Panza' in sidekicks.values())
print('Batman' in sidekicks)
print('Han Solo' not in sidekicks)
print('' not in sidekicks)
"""

########################################################################################################################
"""
Operations with dictionary
Keys and values

What's the difference between keys, values, items dictionary methods? Match the methods with their possible output based
 on the following dictionary:
random_dict = {'a': 20, 'b': 40, 'c': 60, 'd': 80, 'e': 100}
Note that there's one extra option.

Solution
random_dict.keys()      ['a', 'b', 'c', 'd', 'e']
random_dict.values()    [20, 40, 60, 80, 100]
random_dict.items()     [('a', 20), ('b', 40), ('c', 60), ('d', 80), ('e', 100)]
Extra option            ['20', '40', '60', '80', '100']

X       False
        True
        KeyError
        None
"""

########################################################################################################################
"""
Operations with dictionary
Choose the result

When dealing with dict.keys(), you can use some operations for combining them. For instance, & will denote the 
intersection of dictionary keys even if their values are different. Have a look at the code below. What is the right 
output?

tim_toys = {'teddy bear': 3, 'toy car': 5, 'lion': 7, 'puppy': 5}
tom_toys = {'doll': 3, 'puppy': 2, 'kitten': 4, 'teddy bear': 3}
print(tim_toys.keys() & tom_toys.keys())

Solution
X       {'teddy bear', 'puppy'}
        {'teddy bear'}
        {'puppy'}
        {'doll', 'toy car'}
"""

########################################################################################################################
"""
Operations with dictionary
Upper and lower

A some_iterable stores words from a sentence. Use dictionary comprehension to create a new dictionary, in which keys 
will be words from some_iterable, written in uppercase letters, and values will be the same words written in lowercase 
letters. Print this dictionary.

 Sample Input 1:
Great loves too must be endured.

 Sample Output 1:
{'GREAT': 'great', 'LOVES': 'loves', 'TOO': 'too', 'MUST': 'must', 'BE': 'be', 'ENDURED.': 'endured.'}

# the list with words from string
# please, do not modify it
some_iterable = input().split()

# use dictionary comprehension to create a new dictionary
"""
# # the list with words from string
# # please, do not modify it
# some_iterable = input().split()
#
# # use dictionary comprehension to create a new dictionary
# dictionary = {key.upper(): key.lower() for key in some_iterable}
# print(dictionary)

########################################################################################################################
"""
Operations with dictionary
Long fruit

We have the following list fruits = ['apple', 'kiwi', 'banana', 'orange', 'apricot']. Choose the correct line of code 
which will form a dictionary like this:
    key: fruit name;
    value: number of letters in the fruit name.
Moreover, the new dictionary should only include names longer than 5 characters.

Solution
X       fruits_dict = {element: len(element) for element in fruits if len(element) > 5}
        fruits_dict = {key: len(element) for key, element in fruits.items() if len(element) > 5}
        fruits_dict = {element: len(element) if len(element) > 5 for element in fruits}
        fruits_dict = {element: len(element) for element in fruits.values() if len(element) > 5}
"""

########################################################################################################################
"""
Operations with dictionary
Dating App

There's a list with info about people who search for a date. For each person, a few parameters are specified: their 
gender, age, hobbies, and city.

potential_dates = [{"name": "Julia", "gender": "female", "age": 29,
                    "hobbies": ["jogging", "music"], "city": "Hamburg"},
                   {"name": "Sasha", "gender": "male", "age": 18,
                    "hobbies": ["rock music", "art"], "city": "Berlin"}, 
                   {"name": "Maria", "gender": "female", "age": 35,
                    "hobbies": ["art"], "city": "Berlin"},
                   {"name": "Daniel", "gender": "non-conforming", "age": 50,
                    "hobbies": ["boxing", "reading", "art"], "city": "Berlin"}, 
                   {"name": "John", "gender": "male", "age": 41,
                    "hobbies": ["reading", "alpinism", "museums"], "city": "Munich"}]

Help a new user write a function that selects from the given list people older than 30, interested in art, and living 
in Berlin. The function should return their names as a string, separated by a comma and a space, e.g. "Maria, Daniel" 
for the example list above.
"""
# potential_dates = [{"name": "Julia", "gender": "female", "age": 29,
#                     "hobbies": ["jogging", "music"], "city": "Hamburg"},
#                    {"name": "Sasha", "gender": "male", "age": 18,
#                     "hobbies": ["rock music", "art"], "city": "Berlin"},
#                    {"name": "Maria", "gender": "female", "age": 35,
#                     "hobbies": ["art"], "city": "Berlin"},
#                    {"name": "Daniel", "gender": "non-conforming", "age": 50,
#                     "hobbies": ["boxing", "reading", "art"], "city": "Berlin"},
#                    {"name": "John", "gender": "male", "age": 41,
#                     "hobbies": ["reading", "alpinism", "museums"], "city": "Munich"}]
#
#
# def select_dates(potential_dates):
#     result = list()
#     for member in potential_dates:
#         if member["age"] > 30:
#             if "art" in member["hobbies"]:
#                 if member["city"] == "Berlin":
#                     result.append(member["name"])
#     return ", ".join(result)
#
#
# print(select_dates(potential_dates))

########################################################################################################################
"""
Operations with dictionary
Extreme Points

Write a program that prints keys for the minimum and maximum values of the dictionary.

You are supposed to work with the variable test_dict which is a dictionary.

Sample output for the dictionary {"a": 43, "b": 1233, "c": 8} should be as follows:

min: c
max: b
"""
# import json
#
# # The following line creates a dictionary from the input. Do not modify it, please
# test_dict = json.loads(input())
#
# # Work with the 'test_dict'
# # Solution 1
# max_value = 0
# min_value = 1000000
# # for key, value in test_dict.items():
# #     if value >= max_value:
# #         max_value = value
# #         max_key = key
# #     elif value < min_value:
# #         min_value = value
# #         min_key = key
# # print(f"min: {min_key}")
# # print(f"max: {max_key}")
# # # Solution 2
# print(f' min: {min(test_dict, key=test_dict.get)}')
# print(f' max: {max(test_dict, key=test_dict.get)}')


########################################################################################################################
"""
Operations with dictionary
Test result for capitals

What will be the result of the membership test in the last line of the code?
capitals = {'China': 'Beijing', 'United Kingdom': 'London', 'Nigeria': 'Abuja', 'Nauru': ''}
print('' in capitals)

Solution
"""

########################################################################################################################
"""
Set
Easy calculations

Let's suppose we have 13 cats (and a very big house certainly), 3 dogs, 2 turtles, and a lizard. Then we decide to mix 
this gorgeous company up a bit and buy a parrot. Question: if we use Python for making a set of these pets 
(not their names), how many elements will the new set consist of?

>> 5
"""

########################################################################################################################
"""
Set
Format checks

Choose the right variant:
The set's elements are...

    ordered and immutable.
    none of the variants.
    ordered and mutable.
X   unordered and hashable. 
"""

########################################################################################################################
"""
Set
Letters

Write a program that calculates how many distinct letters the input word has. The word is stored in the variable word. 
Print out the result you'll get.
Sample Input 1: 
mississippi
Sample Output 1: 
4
"""
# word = input()  # the input word
# print(len(set(word)))

########################################################################################################################
"""
Set
Keeping up with methods

Fill in the missing word.
To remove all elements from the set you need to use...method.
X   clear()
    remove()
    delete()
    discard()
"""

########################################################################################################################
"""
Set
True or false?

What will be the output of the following piece of code? Don't use the Python console.
set1 = {'oatmeal', 'millet', 'rice', 'semolina', 'buckwheat'}
set2 = {'oatmeal', 'ric', 'semolina', 'millet', 'buckwheat'}
print(set1 == set2)

Solution
> False
"""

########################################################################################################################
"""
Set
Frozenset constraints

What operations can NOT be performed on a frozenset?
add() method
discard() method
len() function
get all elements using a for loop
clear() method

Solution
X   add() method
X   discard() method
    len() function
    get all elements using a for loop
X   clear() method
"""

########################################################################################################################
"""
Set
Mystery set

There is an unknown set of objects named mystery_set which has been predefined. You don't have access to this set.
Write a program that deletes the input string (stored in the variable string) from the mystery_set. It is NOT guaranteed
that the string is an element of the mystery_set. Don't forget to take that into account!
Don't print anything!
"""
# # mystery_set has been defined
# string = input()
#
# # delete string from mystery_set
# mystery_set.discard(string)

########################################################################################################################
"""
Set
Counting unique

Imagine you have information about subjects three students study. It could be in the following format:

Belov = ["Physics", "Math", "Russian"]
Smith = ["Math", "Geometry", "English"]
Sarada = ["Japanese", "Math", "Physics"]

The subjects can be the same or different. Your task is to find the number of unique subjects. For example, in the lists
above we have 6 different subjects: Physics, Math, Russian, Geometry, English, and Japanese, so you should print 6 as 
the answer.

In this task, you're given the variables to work with: Belov, Smith, and Sarada.
"""
# Belov = ["Physics", "Math", "Russian"]
# Smith = ["Math", "Geometry", "English"]
# Sarada = ["Japanese", "Math", "Physics"]
#
# print(len(set(Belov + Smith + Sarada)))

########################################################################################################################
"""
Set
Memory test
Today you are assisting in a psychological experiment. To test short-term memory, a researcher gives a set of numbers 
to each volunteer and asks to repeat all of them. The order does not matter and repeats are allowed.
The input has been read into two variables for you.
If the volunteer remembers all numbers correctly, print True, otherwise, you should output False.
"""
# numbers = input().split()
# answers = input().split()
# print(set(answers) == set(numbers))

########################################################################################################################
"""
Set operations
Match operators

Match the names of the methods with the corresponding operators.
intersection    &
union           -
update          |=
difference      |

Solution        
intersection    &
union           |
update          |=
difference      -
"""

########################################################################################################################
"""
Set operations
Rivers

Some states of the USA share their names with rivers. We have defined two variables with respective place names.
Print out a new set with river names that don't overlap with given states.

 Sample Input 1:
Alabama Missouri Mississippi
Georgia Alaska Missouri

Sample Output 1:
{'Mississippi', 'Alabama'}
"""

# # work with these variables
# rivers = set(input().split())
# states = set(input().split())
# print(rivers - states)

########################################################################################################################
"""
Set operations
A bunch of wizards

You are given several sets with names of students in different classes. Output the set containing names of all the 
students.
NB! You don't need to read the input; the variables containing the input data are already created for you. 

 Sample Input 1:
Potter Weasley
Lovegood Corner
Malfoy Goyle
Bones Macmillan

Sample Output 1:
{'Potter', 'Weasley', 'Lovegood', 'Corner', 'Malfoy', 'Goyle', 'Bones', 'Macmillan'}
"""
# gryffindor = set(input().split())
# ravenclaw = set(input().split())
# slytherin = set(input().split())
# hufflepuff = set(input().split())
# print(gryffindor | ravenclaw | slytherin | hufflepuff)

########################################################################################################################
"""
Set operations
Operation name

What is the name of the operation that returns a new set filled with objects that only one of the original sets had?
    union
    intersection_update
X   difference
    difference_update
"""

########################################################################################################################
"""
Set operations
Tourists

Eugene and Rose decided to travel abroad. For convenience, they will go to a country where one of them has already been,
but not where they both have.
Output a set of potential countries for them.

 Sample Input 1:
Greece Netherlands Colombia UK
Italy UK Russia Greece Canada

Sample Output 1:
{'Russia', 'Italy', 'Colombia', 'Netherlands', 'Canada'}
"""
# # work with these variables
# eugene = set(input().split())
# rose = set(input().split())
# print(rose.symmetric_difference(eugene))

########################################################################################################################
"""
Set operations
Planet sets

Using the three given sets, write a code that creates a set containing the elements that all of the original sets have 
in common. Output this resulting set.

NB! You don't need to read the input; the variables containing the input data are already created for you. 

 Sample Input 1:
Jupiter Saturn Mars
Earth Mars Venus
Mars Pluto Uranus

Sample Output 1:
{'Mars'}

"""
# set_1 = set(input().split())
# set_2 = set(input().split())
# set_3 = set(input().split())
#
# print(set_1 & set_2 & set_3)

########################################################################################################################
"""
Set operations
A set and a string

Write down the command that should be put instead of ... to get a set of symbols that the set a and the string b both 
contain.

a = set("my code is brOKen")
b = "i'm not OK with that"
c = ...

Do NOT convert b to a set in your solution, it should remain a string.
"""
# a = set("my code is brOKen")
# b = "i'm not OK with that"
# c = a.intersection(b)
# print(c)

########################################################################################################################
"""
Set operations
Talents

Imagine you have two lists of names: people from the first list play the violin, while people from the second one happen to speak German.
Print a set with names of those who can do both.

 Sample Input 1:
Vanessa-Mae, David Garrett
Rosamond Mayhan, Malik Kopf, David Garrett, Cortez Mestas, Barbara Miller, Elease Knudson

Sample Output 1:
{'David Garrett'}
"""
# # work with these variables
# violinists = set(input().split(', '))
# german_speakers = set(input().split(', '))
# print(violinists & german_speakers)

########################################################################################################################
"""
Set operations
Hack the Pentagon

There's a hidden variable with top secret data. You can't see it, and all you know about it is that it's called 
pentagon_passwords and that it is a container with sets. Try to get the intersection of these sets and output the number
of elements in it. 
"""
# pentagon_intersection = set.intersection(*pentagon_passwords)
# print(len(pentagon_intersection))

########################################################################################################################
"""
Set operations
Operator and method

What is the difference between using an operator and a set operation method?
X   The difference is in the syntax.
    Only operators can work with non-set arguments
    Only operators can be used to update an existing set
X   Only methods can accept a bunch of sets at once, in a certain way
"""

########################################################################################################################
"""
How to choose a collection to use
Dictionary and Slicing

Assume you have a dictionary of children's names and toys that belong to them.

names_toys = {"Tom": "car", "Suzy": "teddy bear", "Sam": "boat", "Lucy": "ball"}

You want to know the first two entries of this dictionary ({"Tom": "car", "Suzy": "teddy bear"}). To do so, you must 
slice the dictionary. Is this correct?

    Yes
X   No
    Depends on which Python version I'm using
"""

########################################################################################################################
"""
How to choose a collection to use
RGB colours

According to the RGB model, red, green, and blue can be added together in various ways to create different colors. 
Assume you have a list of the following colors and their corresponding RGB values:
	        R 	    G 	    B
Purple 	    160 	32 	    255
Light Blue 	80 	    208 	255
Yellow 	    255 	224 	32

Write a program that takes the name of one of these colors as input and returns its appropriate RGB values. 

Sample Input 1:
Purple

Sample Output 1:
(160, 32, 255)
"""
# color = input()
# colors = {"purple": (160, 32, 255), "light blue": (80, 208, 255), "yellow": (255, 224, 32)}
# print(colors[color])

########################################################################################################################
"""
How to choose a collection to use
Storage type

Match collections and storage types.
Report a typo
Match the items from left and right columns
List        row/column
Set         row
Dictionary  key-value

Solution
List        row/column
Set         row
Dictionary  key-value
"""

########################################################################################################################
"""
How to choose a collection to use
Immutability

What collection isn't mutable?
Report a typo
Select one option from the list
list
tuple
dictionary
set

Solution
    list
X   tuple
    dictionary
    set
"""

########################################################################################################################
"""
How to choose a collection to use
Duplicates

Select one or more options from the list
Tuples
Sets
Dictionary keys
Dictionary values
Lists

Solution
    Tuples
X   Sets
X   Dictionary keys
    Dictionary values
    Lists
"""

########################################################################################################################
"""
How to choose a collection to use
Grocery shopping

Every Sunday Jane goes to a local market to buy groceries for the upcoming week. When she makes her shopping list, she 
never puts everything down at once: she keeps track of what she wants to buy during the day. On Sunday morning her list 
looked like this:

to_buy = ['carrots', 'carrots', 'bread', 'tomatoes', 'onions', 'apples', 'tomatoes', 'carrots', 
          'tomatoes', 'onions', 'onions', 'onions', 'bread', 'milk', 'bread', 'apples']

Write a program that prints what she wants to buy and how many times she has added it to her shopping list, like in the 
example. The order of the items should be preserved.
Report a typo

Sample Input 1:

carrots carrots bread tomatoes onions apples tomatoes carrots tomatoes onions onions onions bread milk bread apples

Sample Output 1:

3 carrots
3 bread
3 tomatoes
4 onions
2 apples
1 milk
"""
# shopping_list = input().split()
# to_buy = {}
#
# for item in shopping_list:
#     to_buy.update({item: shopping_list.count(item)})
#
# for item, nb_item in to_buy.items():
#     print(nb_item, item)

########################################################################################################################
"""
How to choose a collection to use
Seasons

You were asked to create a stacked tuple called months (a so-called "tuple of tuples") with the names of months. Each 
tuple contains months belonging to the corresponding season of the year starting with winter months:

months = (('December', 'January', 'February'), 
          ('March', 'April', 'May'), 
          ('June', 'July', 'August'),
          ('September', 'October', 'November'))

The months tuple is already defined in this task, you don't need to create it.

You will get 3 successive months of the year in the input. Your task is to read the input, create a tuple of input 
months and check if it is in the months tuple. If it is so, print True, otherwise, False.
Report a typo

Sample Input 1:

December January February

Sample Output 1:

True
"""
# months = (('December', 'January', 'February'),
#           ('March', 'April', 'May'),
#           ('June', 'July', 'August'),
#           ('September', 'October', 'November'))
#
# three_months = input().split()
# print(tuple(three_months) in months)

########################################################################################################################
"""
How to choose a collection to use
Fill in the gap

Fill in the gap in the following statement:
Only _________ are unordered in Python 3.7.
Solution
Sets
"""

########################################################################################################################
"""
How to choose a collection to use
Matching

List        Store guests' names for a party (it is updated constantly).
Tuple       Check student's attendance (it is immutable).
Dictionary  For each employee, store their name and age, accessed by their ID.
Set         For a group of people, write down in what months they have their birthdays. You are not interested how many 
of them have a birthday in a particular month.

"""

########################################################################################################################
"""
How to choose a collection to use
Remove duplicates

Suppose you have a list of names. Write a program that takes this list as an input, removes all duplicate names, and 
returns them in alphabetical order.

To return the names in alphabetical or numerical order, you can use the following built-in methods: list_name.sort() for
lists and sorted(variable_name) for sets, tuples or lists. The difference between these two methods is that sort() 
modifies the original list, and sorted() creates a sorted copy of a collection.
"""
names = input().split()
for name in sorted(set(names)):
    print(name)


########################################################################################################################
"""
Work on project. Stage 1/4 Mini-calculator
Mini-calculator

Description

People learn new things in one way or another. Learning sometimes means that you need to check your comprehension by 
taking a test. It also requires a person (or a program) to check your answers. You may have been in a situation when 
you think that you have solved the task correctly, but your professor has a different (sometimes wrong!) answer. It 
happens; everybody makes mistakes.
Not our application though. It should calculate the solution in a very precise manner. We need to make a simple 
calculator that can evaluate expressions like a + b, a - b, or a * b. We will leave the division aside for now.
Objectives

    A user inputs a line that looks like a simple mathematical operation.
    The application should print the result of the operation.

Examples

The greater-than symbol followed by a space (> ) represents the user input. Note that it's not part of the input.
"""

# entry = input().split()
#
# if entry[1] == '+':
#     print(int(entry[0]) + int(entry[2]))
# elif entry[1] == '-':
#     print(int(entry[0]) - int(entry[2]))
# elif entry[1] == '*':
#     print(int(entry[0]) * int(entry[2]))

########################################################################################################################
"""
Declaring a function 
Fahrenheit 

Convert the temperature from Fahrenheit to Celsius in the function below. You can use this formula:

C∘=(F∘−32)×59C^\circ = (F^\circ - 32)\times \frac{5}{9}C∘=(F∘−32)×95​

Round the returned result to 3 decimal places.

You don't have to handle input, just implement the function below.

Also, make sure your function returns the value. Please do NOT print anything.
"""

# def fahrenheit_to_celsius(fahrenheit):
#     return round((fahrenheit - 32) * 5 / 9, 3)
#
#
# print(fahrenheit_to_celsius(451))

########################################################################################################################
"""
Declaring a function 
Make the function work

The function closest_higher_mod_5 takes exactly one integer argument x and returns the smallest integer y such that:

    y is greater than or equal to x,
    y is divisible by 5.

Correct the last line of the code below to make the function work.

def closest_higher_mod_5(x):
    remainder = x % 5
    if remainder == 0:
        return x
    return "I don't know :("
"""

# def closest_higher_mod_5(x):
#     remainder = x % 5
#     while remainder > 0:
#         x += 1
#         remainder = x % 5
#     else:
#         return x if remainder == 0 else "I don't know :("
#
#
# print(closest_higher_mod_5(43))

########################################################################################################################
"""
Declaring a function 
Captain

Define a function that will add the word "captain" before the name of a person. The function should be named 
captain_adder, take one argument name, and print the string, i.e. it doesn't have to return anything. You don't need to 
call your function, only implement it. Also, you don't need to take input, use only the name variable that should be an 
argument of the function. 
"""

# def captain_adder(name):
#     print("captain ", name)
#
# captain_adder("Julien")

########################################################################################################################
"""
Declaring a function 
Prevent cheating

On Hyperskill, we sometimes need to check that the student's code doesn't use a particular method. For instance, when 
the task is to manually implement the factorial computation instead of using the factorial() function from the math 
module. How could we do that?

It is straightforward. We can override math.factorial() to do something else — raise an error or print a message to the 
student. To do so, first, we define a function that does what we want (for example, prints a message) with the same 
arguments (since we want to disguise our new function as the original one). Then, outside this new function body, we 
need to assign it to math.factorial this way: math.factorial = new_math_factorial (pay attention to the absence of 
parentheses). Now, if the student tries to use math.factorial, our custom function will be raised. It happens because 
when multiple functions with the same name exist, the later one always overrides the prior.

In this task, you are a content creator at Hyperskill. Implement the example with the factorial yourself. Your program 
should override math.factorial() and make it print the message Don't cheat! instead of calculating a factorial. 
"""
# import math
#
# # your code here
#
#
# def new_math_factorial(number):
#     print("Don't cheat!")
#
#
# math.factorial = new_math_factorial
#
# # don't delete this line, please
# math.factorial(23)

########################################################################################################################
"""
Else statement
Spellchecker

Whoa! This problem requires knowledge of list collection type. If you're feeling up to the challenge, brace yourself, 
and good luck! Otherwise, you can skip it for now and return any time later.
Write a simple spellchecker that tells you if the word is spelled correctly. Use the dictionary in the code below: it 
contains the list of all correctly written words.

The input format:
A single line with the "word".

The output format:
If the word is spelled correctly write Correct, otherwise, Incorrect. 
dictionary = ["aa", "abab", "aac", "ba", "bac", "baba", "cac", "caac"]
"""
# dictionary = ["aa", "abab", "aac", "ba", "bac", "baba", "cac", "caac"]
#
# word = input()
# Solution1:
# if word in dictionary:
#     print("Correct")
# else:
#     print("Incorrect")

# Solution2:
# print("Correct" if word in dictionary else "Incorrect")

########################################################################################################################
"""
Else statement
The Maximum
Here is a program which finds the maximum of two numbers. Guess what it will print for the variables a = 9, b = 9 and 
select the line of the code that will give such a result.

if a > b:
    print(a)
else:
    print(b)
"""

# if a > b:
#     print(a)
# else:
#     print(b)

########################################################################################################################
"""
Else statement
Minimum and maximum 

Imagine that your friend asked you to write a program that finds the minimum and the maximum.
Write a program that receives two integers as its input, each number goes on a new line. The output should show:
    The biggest number in the first line
    The smallest number in the second line.
Note that your friend might insert identical numbers! In this case, just output both given numbers on separate lines.
"""
# first_number = int(input())
# second_number = int(input())
#
# if first_number > second_number:
#     print(first_number)
#     print(second_number)
# elif second_number > first_number:
#     print(second_number)
#     print(first_number)
# else:
#     print(first_number)
#     print(second_number)

########################################################################################################################
"""
Else statement
Triangle or not triangle

Read three angles given on separate input lines and check whether they form a triangle. Print the answer in the 
following format: "The triangle is valid!" or "The triangle is not valid!".
"""
# sum_triangle_angles = 180
# angles = []
# for _ in range(0, 3):
#     angles.append(int(input()))
#
# if sum(angles) == sum_triangle_angles:
#     print("The triangle is valid!")
# else:
#     print("The triangle is not valid!")

########################################################################################################################
"""
Else statement
Opposite Sign

Read an integer as input and print it with the opposite sign.
"""
# number = int(input())
# print(-number)

########################################################################################################################
"""
While loop
Rich man's world

When a bank has financial problems the government can return a client's deposit if it is less than 700,000. The 
interest rate for a particular deposit is 7.1% a year. The percentages are paid to the same deposit at the end of the 
year and a new value of the interest is calculated.
Find out how many years it will take for the sum of the deposit to exceed the value protected by the government.

The input format:
The initial sum of the deposit. It is guaranteed that the value will be between 50,000 and 700,000.

The output format:
The number of years.
"""
# initial_deposit = int(input())
# ir = 7.1
# allowed_returned_deposit = 700000
# new_account_value = 0
# year = 0
# new_account_value += initial_deposit
#
# while new_account_value < allowed_returned_deposit:
#     new_account_value += new_account_value * ir / 100
#     year += 1
#
# print(year)

########################################################################################################################
"""
While loop
Favor

Carl asks you to count the sum of the first k natural numbers. Read k from the input, then add up numbers from 1 to k 
and print your answer.
"""

# k = int(input())
# print(sum(range(0, k + 1)))

########################################################################################################################
"""
While loop
Half-life

In nuclear physics, the half-life is used to describe the rate with which elements undergo radioactive decay. More 
precisely, it is the time required for an element to reduce in half.

Let's take an isotope of Radium (Ra) called radium-223. Its half-life is about 12 days: this means that every 12 days 
the number of atoms reduces in half.

Your program should:

    read the initial and the final quantity of atoms from the input.
    count how many complete half-life periods it would take for the initial number of atoms of radium-223 to become 
    equal to or go below the final quantity. Note that the number of half-life periods should be an integer, not a float!
    multiply the number of half-life periods by 12 to convert the number of half-life periods to days.
    output the resulting number of days that it takes for the initial number of atoms to go below the final number.

For example, the initial number of atoms is 4 and the resulting quantity is 3. After the first half-life period, the 
initial number will reduce to 2 atoms, which is below 3. Then, we take the number of half-life periods that have passed 
(1) and multiply it by the number of days that every half-life period takes (12). The output will be 12.
The input format:

The first line: the initial quantity of atoms (from 2 to 1,000,000).
The second line: the final quantity of atoms.

The output format:

The number of days that it would take for radium-223 to go from the initial quantity of atoms to or below the final
quantity of atoms.
An example:
The initial quantity is 8, the final quantity is 3. Your program output should be number 24.
"""

# initial_quantity = int(input())
# final_quantity = int(input())
# half_period = 12
# number_of_half_period = 0
#
# while initial_quantity >= final_quantity:
#     number_of_half_period += 1
#     initial_quantity -= int(initial_quantity / 2)
#
# print(number_of_half_period * 12)

########################################################################################################################
